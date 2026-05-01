"""
Generadores de reportes PDF y Excel para el modulo de informes.

Este modulo encapsula la construccion de archivos exportables para la
epica de estadisticas. Mantiene separada la logica de formato respecto
de los servicios de negocio y entrega objetos reutilizables por la capa
API en la tarea siguiente.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any, Optional, Sequence

from django.conf import settings
from django.core.files.base import ContentFile
from django.db.models import DecimalField, ExpressionWrapper, F, Min
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from cliente.models import Cliente
from core.exceptions import InformeError
from inventario.models import Producto
from ventas.models import DetalleVenta, Venta

from .services import (
    CierreCajaService,
    ReporteEstadisticasService,
    _local_day_start_utc,
    _next_local_day_start_utc,
)


ZERO = Decimal('0.00')
QUANTIZER = Decimal('0.01')
PDF_MIME_TYPE = 'application/pdf'
EXCEL_MIME_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
)
BUSINESS_NAME = 'Mallor'
PDF_ACCENT = colors.HexColor('#1F4E79')
PDF_ACCENT_LIGHT = colors.HexColor('#D9E8F5')
PDF_BORDER = colors.HexColor('#B8CCE4')
PDF_TEXT = colors.HexColor('#203040')
THIN_SIDE = Side(style='thin', color='D9E2F3')
EXCEL_HEADER_FILL = PatternFill(
    fill_type='solid',
    start_color='1F4E79',
    end_color='1F4E79',
)
EXCEL_SUBHEADER_FILL = PatternFill(
    fill_type='solid',
    start_color='D9E8F5',
    end_color='D9E8F5',
)
EXCEL_ALT_FILL = PatternFill(
    fill_type='solid',
    start_color='F6FAFD',
    end_color='F6FAFD',
)
EXCEL_ALERT_FILL = PatternFill(
    fill_type='solid',
    start_color='FDE9E7',
    end_color='FDE9E7',
)
EXCEL_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)
EXCEL_TITLE_FONT = Font(name='Calibri', size=14, bold=True, color='1F1F1F')
EXCEL_HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
EXCEL_SUBHEADER_FONT = Font(
    name='Calibri',
    size=11,
    bold=True,
    color='1F1F1F',
)
EXCEL_BODY_FONT = Font(name='Calibri', size=10, color='1F1F1F')
EXCEL_META_FONT = Font(name='Calibri', size=10, italic=True, color='595959')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_WRAP = Alignment(
    horizontal='left',
    vertical='center',
    wrap_text=True,
)
NUMBER_FORMAT_CURRENCY = '#,##0.00'
NUMBER_FORMAT_INTEGER = '#,##0'
NUMBER_FORMAT_QUANTITY = '#,##0.00'
NUMBER_FORMAT_PERCENT = '0.00%'
NUMBER_FORMAT_DATE = 'YYYY-MM-DD'
DEFAULT_LOGO_CANDIDATES = (
    'logo.png',
    'logo.jpg',
    'logo.jpeg',
    'logo.webp',
    'static/logo.png',
    'static/img/logo.png',
    'media/logo.png',
)


@dataclass(frozen=True)
class GeneratedReportFile:
    """
    Resultado serializable de una exportacion de informe.

    La capa API puede adjuntar este resultado a un modelo `Informe`,
    escribirlo en un `FileField` o retornarlo como descarga HTTP.
    """

    filename: str
    content: bytes
    content_type: str

    @property
    def size(self) -> int:
        """Retorna el tamano del archivo en bytes."""
        return len(self.content)

    def to_content_file(self) -> ContentFile:
        """Convierte el contenido a `ContentFile` de Django."""
        return ContentFile(self.content, name=self.filename)


class BaseReportGenerator:
    """Utilidades compartidas por los generadores PDF y Excel."""

    def __init__(self) -> None:
        self.styles = getSampleStyleSheet()
        self.styles.add(
            ParagraphStyle(
                name='MallorTitle',
                parent=self.styles['Heading1'],
                fontName='Helvetica-Bold',
                fontSize=18,
                leading=22,
                textColor=PDF_ACCENT,
                spaceAfter=8,
            ),
        )
        self.styles.add(
            ParagraphStyle(
                name='MallorSection',
                parent=self.styles['Heading2'],
                fontName='Helvetica-Bold',
                fontSize=11,
                leading=14,
                textColor=PDF_ACCENT,
                spaceBefore=8,
                spaceAfter=6,
            ),
        )
        self.styles.add(
            ParagraphStyle(
                name='MallorBody',
                parent=self.styles['BodyText'],
                fontName='Helvetica',
                fontSize=9,
                leading=12,
                textColor=PDF_TEXT,
            ),
        )
        self.styles.add(
            ParagraphStyle(
                name='MallorMeta',
                parent=self.styles['BodyText'],
                fontName='Helvetica',
                fontSize=8,
                leading=10,
                textColor=colors.HexColor('#637381'),
            ),
        )

    @staticmethod
    def _parse_date(value: Any, field_name: str) -> date:
        if isinstance(value, datetime):
            if timezone.is_aware(value):
                return timezone.localtime(value).date()
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, str):
            try:
                return date.fromisoformat(value)
            except ValueError as exc:
                raise InformeError(
                    _(
                        'El campo %(campo)s debe tener una fecha valida.'
                    ) % {
                        'campo': field_name,
                    },
                    code='fecha_invalida',
                ) from exc

        raise InformeError(
            _(
                'El campo %(campo)s debe ser una fecha valida.'
            ) % {
                'campo': field_name,
            },
            code='fecha_invalida',
        )

    @staticmethod
    def _quantize(value: Any) -> Decimal:
        if value in (None, '', [], {}):
            return ZERO

        if isinstance(value, Decimal):
            decimal_value = value
        else:
            try:
                decimal_value = Decimal(str(value))
            except (InvalidOperation, TypeError, ValueError):
                decimal_value = ZERO

        return decimal_value.quantize(QUANTIZER)

    @staticmethod
    def _format_currency(value: Any) -> str:
        return f"${BaseReportGenerator._quantize(value):,.2f}"

    @staticmethod
    def _format_number(value: Any) -> str:
        return f"{BaseReportGenerator._quantize(value):,.2f}"

    @staticmethod
    def _format_date(value: Optional[date]) -> str:
        if value is None:
            return ''
        return value.strftime('%Y-%m-%d')

    @staticmethod
    def _format_datetime(value: Optional[datetime]) -> str:
        if value is None:
            return ''

        if timezone.is_aware(value):
            value = timezone.localtime(value)

        return value.strftime('%Y-%m-%d %H:%M')

    @staticmethod
    def _build_filename(
        prefix: str,
        *parts: Any,
        extension: str,
    ) -> str:
        sanitized_parts = []

        for part in parts:
            text = str(part).strip().replace('/', '-').replace('\\', '-')
            text = text.replace(' ', '-').replace(':', '-')
            if text:
                sanitized_parts.append(text.lower())

        base_name = '-'.join([prefix, *sanitized_parts]).strip('-')
        return f'{base_name}.{extension}'

    @staticmethod
    def _resolve_logo_path() -> Optional[Path]:
        base_dir = Path(settings.BASE_DIR)

        for candidate in DEFAULT_LOGO_CANDIDATES:
            candidate_path = base_dir / candidate
            if candidate_path.exists() and candidate_path.is_file():
                return candidate_path

        return None

    @staticmethod
    def _ventas_queryset_periodo(
        fecha_inicio: date,
        fecha_fin: date,
    ):
        if fecha_inicio > fecha_fin:
            raise InformeError(
                _(
                    'La fecha de inicio no puede ser posterior a la fecha '
                    'de fin.'
                ),
                code='rango_fechas_invalido',
            )

        return Venta.objects.filter(
            estado=Venta.Estado.TERMINADA,
            fecha_venta__gte=_local_day_start_utc(fecha_inicio),
            fecha_venta__lt=_next_local_day_start_utc(fecha_fin),
        )

    @staticmethod
    def _detalles_queryset_periodo(
        fecha_inicio: date,
        fecha_fin: date,
    ):
        return DetalleVenta.objects.filter(
            venta__in=BaseReportGenerator._ventas_queryset_periodo(
                fecha_inicio,
                fecha_fin,
            ),
        )

    @staticmethod
    def _historical_sales_range() -> tuple[date, date]:
        fecha_minima = Venta.objects.filter(
            estado=Venta.Estado.TERMINADA,
        ).aggregate(
            fecha_minima=Min('fecha_venta'),
        )['fecha_minima']
        fecha_fin = timezone.localdate()

        if fecha_minima is None:
            return fecha_fin, fecha_fin

        if timezone.is_aware(fecha_minima):
            fecha_inicio = timezone.localtime(fecha_minima).date()
        else:
            fecha_inicio = fecha_minima.date()

        return fecha_inicio, fecha_fin

    @staticmethod
    def _truncate_items(
        items: Sequence[dict[str, Any]],
        limit: int = 10,
    ) -> list[dict[str, Any]]:
        if len(items) <= limit:
            return list(items)
        return list(items[:limit])

    @staticmethod
    def _cliente_nombre(venta: Venta) -> str:
        if venta.cliente_id is None:
            return 'Consumidor Final'
        return venta.cliente.get_nombre_completo()

    @staticmethod
    def _dias_vencidos(venta: Venta) -> int:
        fecha_venta = timezone.localtime(venta.fecha_venta).date()
        dias_plazo = venta.cliente.dias_plazo if venta.cliente else 0
        vencimiento = fecha_venta + timedelta(days=dias_plazo)
        return (timezone.localdate() - vencimiento).days

    @staticmethod
    def _bucket_cartera(venta: Venta) -> str:
        dias_vencidos = BaseReportGenerator._dias_vencidos(venta)

        if dias_vencidos <= 0:
            return 'AL_DIA'
        if dias_vencidos <= 30:
            return '1_30'
        if dias_vencidos <= 60:
            return '31_60'
        if dias_vencidos <= 90:
            return '61_90'
        return '91_MAS'

    @staticmethod
    def _chart_points(
        rows: Sequence[dict[str, Any]],
        label_key: str,
        value_key: str,
        limit: int = 12,
    ) -> tuple[list[str], list[float]]:
        if not rows:
            return [], []

        if len(rows) <= limit:
            selected = list(rows)
        else:
            step = max(1, len(rows) // limit)
            selected = list(rows[::step])[:limit]

        labels = [str(item.get(label_key, '')) for item in selected]
        values = [float(item.get(value_key, 0.0) or 0.0) for item in selected]
        return labels, values

    @staticmethod
    def _build_pdf_result(filename: str, buffer: BytesIO) -> GeneratedReportFile:
        return GeneratedReportFile(
            filename=filename,
            content=buffer.getvalue(),
            content_type=PDF_MIME_TYPE,
        )

    @staticmethod
    def _build_excel_result(
        filename: str,
        buffer: BytesIO,
    ) -> GeneratedReportFile:
        return GeneratedReportFile(
            filename=filename,
            content=buffer.getvalue(),
            content_type=EXCEL_MIME_TYPE,
        )


class PDFReportGenerator(BaseReportGenerator):
    """Generador especializado de reportes PDF."""

    def _header_logo(self) -> Any:
        logo_path = self._resolve_logo_path()

        if logo_path is not None:
            image = Image(str(logo_path), width=2.2 * cm, height=2.2 * cm)
            image.hAlign = 'LEFT'
            return image

        drawing = Drawing(80, 32)
        drawing.add(
            Rect(
                0,
                0,
                28,
                28,
                fillColor=PDF_ACCENT,
                strokeColor=PDF_ACCENT,
                radius=4,
            ),
        )
        drawing.add(
            String(
                8,
                7,
                'M',
                fontName='Helvetica-Bold',
                fontSize=16,
                fillColor=colors.white,
            ),
        )
        drawing.add(
            String(
                36,
                11,
                BUSINESS_NAME,
                fontName='Helvetica-Bold',
                fontSize=14,
                fillColor=PDF_ACCENT,
            ),
        )
        return drawing

    def _header_block(
        self,
        title: str,
        filters: Sequence[str],
    ) -> Table:
        logo = self._header_logo()
        metadata = [
            Paragraph(title, self.styles['MallorTitle']),
            Paragraph(
                f'Fecha de generacion: '
                f'{self._format_datetime(timezone.now())}',
                self.styles['MallorMeta'],
            ),
        ]

        for line in filters:
            metadata.append(
                Paragraph(line, self.styles['MallorMeta']),
            )

        data = [[logo, metadata]]
        table = Table(data, colWidths=[3.2 * cm, 13.8 * cm])
        table.setStyle(
            TableStyle(
                [
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 0),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                    ('TOPPADDING', (0, 0), (-1, -1), 0),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ],
            ),
        )
        return table

    def _section_title(self, text: str) -> Paragraph:
        return Paragraph(text, self.styles['MallorSection'])

    def _body(self, text: str) -> Paragraph:
        return Paragraph(text, self.styles['MallorBody'])

    def _metric_table(self, rows: Sequence[Sequence[str]]) -> Table:
        table = Table(rows, colWidths=[7.2 * cm, 4.0 * cm, 4.0 * cm])
        table.setStyle(
            TableStyle(
                [
                    ('BACKGROUND', (0, 0), (-1, 0), PDF_ACCENT),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 0.5, PDF_BORDER),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                        colors.white,
                        PDF_ACCENT_LIGHT,
                    ]),
                    ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ],
            ),
        )
        return table

    def _data_table(
        self,
        headers: Sequence[str],
        rows: Sequence[Sequence[Any]],
        col_widths: Optional[Sequence[float]] = None,
    ) -> Table:
        data = [list(headers), *[list(row) for row in rows]]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ('BACKGROUND', (0, 0), (-1, 0), PDF_ACCENT),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 0.5, PDF_BORDER),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [
                        colors.white,
                        PDF_ACCENT_LIGHT,
                    ]),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ],
            ),
        )
        return table

    def _bar_chart(
        self,
        title: str,
        labels: Sequence[str],
        values: Sequence[float],
    ) -> Optional[Drawing]:
        if not labels or not values:
            return None

        drawing = Drawing(460, 190)
        drawing.add(
            String(
                0,
                176,
                title,
                fontName='Helvetica-Bold',
                fontSize=10,
                fillColor=PDF_ACCENT,
            ),
        )
        chart = VerticalBarChart()
        chart.x = 40
        chart.y = 28
        chart.height = 128
        chart.width = 390
        chart.data = [list(values)]
        chart.categoryAxis.categoryNames = list(labels)
        chart.categoryAxis.labels.boxAnchor = 'ne'
        chart.categoryAxis.labels.angle = 30
        chart.categoryAxis.labels.dy = -2
        chart.categoryAxis.labels.fontSize = 7
        chart.valueAxis.valueMin = 0
        chart.bars[0].fillColor = PDF_ACCENT
        chart.bars[0].strokeColor = PDF_ACCENT
        chart.strokeColor = PDF_BORDER
        drawing.add(chart)
        return drawing

    def _pie_chart(
        self,
        title: str,
        labels: Sequence[str],
        values: Sequence[float],
    ) -> Optional[Drawing]:
        if not labels or not values:
            return None

        drawing = Drawing(460, 220)
        drawing.add(
            String(
                0,
                204,
                title,
                fontName='Helvetica-Bold',
                fontSize=10,
                fillColor=PDF_ACCENT,
            ),
        )
        pie = Pie()
        pie.x = 120
        pie.y = 20
        pie.width = 170
        pie.height = 170
        pie.data = list(values)
        pie.labels = list(labels)
        pie.slices.strokeWidth = 0.5
        pie.sideLabels = True

        palette = [
            colors.HexColor('#1F4E79'),
            colors.HexColor('#4F81BD'),
            colors.HexColor('#9CC2E5'),
            colors.HexColor('#C6E0F5'),
            colors.HexColor('#D9EAD3'),
        ]
        for index, _ in enumerate(values):
            pie.slices[index].fillColor = palette[index % len(palette)]

        drawing.add(pie)
        return drawing

    def _footer(self, canvas, document) -> None:
        canvas.saveState()
        canvas.setStrokeColor(PDF_BORDER)
        canvas.line(
            document.leftMargin,
            1.3 * cm,
            A4[0] - document.rightMargin,
            1.3 * cm,
        )
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(colors.HexColor('#6B7280'))
        canvas.drawString(
            document.leftMargin,
            0.8 * cm,
            f'{BUSINESS_NAME} - Reporte generado automaticamente',
        )
        canvas.drawRightString(
            A4[0] - document.rightMargin,
            0.8 * cm,
            f'Pagina {canvas.getPageNumber()}',
        )
        canvas.restoreState()

    def _build_document(
        self,
        title: str,
        filters: Sequence[str],
        story: list[Any],
        filename: str,
        pagesize=A4,
    ) -> GeneratedReportFile:
        buffer = BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            leftMargin=1.6 * cm,
            rightMargin=1.6 * cm,
            topMargin=1.6 * cm,
            bottomMargin=1.7 * cm,
        )
        elements = [self._header_block(title, filters), Spacer(1, 0.35 * cm)]
        elements.extend(story)
        document.build(
            elements,
            onFirstPage=self._footer,
            onLaterPages=self._footer,
        )
        buffer.seek(0)
        return self._build_pdf_result(filename, buffer)

    def generar_pdf_ventas_periodo(
        self,
        datos: Optional[dict[str, Any]],
        fecha_inicio: Any,
        fecha_fin: Any,
    ) -> GeneratedReportFile:
        """
        Genera el PDF ejecutivo de ventas por periodo.
        """
        fecha_inicio_date = self._parse_date(fecha_inicio, 'fecha_inicio')
        fecha_fin_date = self._parse_date(fecha_fin, 'fecha_fin')
        estadisticas = (
            datos
            if isinstance(datos, dict) and datos
            else ReporteEstadisticasService.estadisticas_ventas_periodo(
                fecha_inicio_date,
                fecha_fin_date,
            )
        )
        categorias = ReporteEstadisticasService.ventas_por_categoria(
            fecha_inicio_date,
            fecha_fin_date,
        )
        metodos = ReporteEstadisticasService.ventas_por_metodo_pago(
            fecha_inicio_date,
            fecha_fin_date,
        )
        top_productos = ReporteEstadisticasService.productos_mas_vendidos(
            fecha_inicio_date,
            fecha_fin_date,
            10,
        )
        ventas = list(
            self._ventas_queryset_periodo(
                fecha_inicio_date,
                fecha_fin_date,
            ).select_related(
                'cliente',
                'usuario_registro',
            ).order_by('-fecha_venta')[:20]
        )
        resumen = estadisticas.get('resumen', {})
        comparacion = estadisticas.get(
            'comparacion_periodo_anterior',
            {},
        )
        story: list[Any] = []
        story.append(self._section_title('Resumen general'))
        story.append(
            self._metric_table(
                [
                    ['Indicador', 'Periodo actual', 'Periodo anterior'],
                    [
                        'Total ventas',
                        self._format_currency(resumen.get('total_ventas')),
                        self._format_currency(
                            comparacion.get(
                                'total_ventas',
                                {},
                            ).get('anterior'),
                        ),
                    ],
                    [
                        'Cantidad ventas',
                        str(resumen.get('cantidad_ventas', 0)),
                        str(
                            comparacion.get(
                                'cantidad_ventas',
                                {},
                            ).get('anterior', 0),
                        ),
                    ],
                    [
                        'Ticket promedio',
                        self._format_currency(
                            resumen.get('ticket_promedio'),
                        ),
                        self._format_currency(
                            comparacion.get(
                                'ticket_promedio',
                                {},
                            ).get('anterior'),
                        ),
                    ],
                ],
            ),
        )
        story.append(Spacer(1, 0.25 * cm))

        series = estadisticas.get('grafico_tendencia', [])
        labels, values = self._chart_points(
            series,
            'fecha',
            'total_ventas',
        )
        chart = self._bar_chart(
            'Tendencia diaria de ventas',
            labels,
            values,
        )
        if chart is not None:
            story.append(chart)
            story.append(Spacer(1, 0.2 * cm))

        story.append(self._section_title('Distribucion por metodo de pago'))
        metodo_rows = [
            [
                item['metodo_pago'],
                self._format_currency(item['total_vendido']),
                f"{item['porcentaje']:.2f}%",
            ]
            for item in metodos.get('distribucion', [])
        ]
        story.append(
            self._data_table(
                ('Metodo', 'Total vendido', 'Participacion'),
                metodo_rows or [['Sin datos', '$0.00', '0.00%']],
                col_widths=[6 * cm, 5 * cm, 4.2 * cm],
            ),
        )
        story.append(Spacer(1, 0.2 * cm))
        metodo_labels, metodo_values = self._chart_points(
            metodos.get('distribucion', []),
            'metodo_pago',
            'total_vendido',
            limit=5,
        )
        metodo_chart = self._pie_chart(
            'Participacion por metodo de pago',
            metodo_labels,
            metodo_values,
        )
        if metodo_chart is not None:
            story.append(metodo_chart)

        story.append(self._section_title('Ventas por categoria'))
        categoria_rows = [
            [
                item['categoria'],
                self._format_currency(item['total_vendido']),
                f"{item['porcentaje']:.2f}%",
            ]
            for item in categorias.get('distribucion', [])
        ]
        story.append(
            self._data_table(
                ('Categoria', 'Total vendido', 'Participacion'),
                categoria_rows or [['Sin datos', '$0.00', '0.00%']],
                col_widths=[7.6 * cm, 4.0 * cm, 3.6 * cm],
            ),
        )
        story.append(Spacer(1, 0.2 * cm))
        categoria_labels, categoria_values = self._chart_points(
            categorias.get('distribucion', []),
            'categoria',
            'total_vendido',
        )
        categoria_chart = self._bar_chart(
            'Categorias con mayor ingreso',
            categoria_labels,
            categoria_values,
        )
        if categoria_chart is not None:
            story.append(categoria_chart)

        story.append(self._section_title('Productos destacados'))
        producto_rows = [
            [
                item['nombre'],
                self._format_number(item['cantidad_vendida']),
                self._format_currency(item['total_vendido']),
                self._format_currency(item['margen_generado']),
            ]
            for item in top_productos.get('resultados', [])
        ]
        story.append(
            self._data_table(
                ('Producto', 'Cantidad', 'Total vendido', 'Margen'),
                producto_rows or [['Sin datos', '0', '$0.00', '$0.00']],
                col_widths=[7.0 * cm, 2.6 * cm, 3.2 * cm, 3.2 * cm],
            ),
        )

        story.append(self._section_title('Ventas recientes del periodo'))
        venta_rows = [
            [
                venta.numero_venta,
                self._format_datetime(venta.fecha_venta),
                self._cliente_nombre(venta),
                venta.get_metodo_pago_display(),
                self._format_currency(venta.total),
            ]
            for venta in ventas
        ]
        story.append(
            self._data_table(
                ('Venta', 'Fecha', 'Cliente', 'Metodo', 'Total'),
                venta_rows or [['Sin datos', '', '', '', '$0.00']],
                col_widths=[2.6 * cm, 3.3 * cm, 6.0 * cm, 2.4 * cm, 2.6 * cm],
            ),
        )

        filename = self._build_filename(
            'ventas-periodo',
            fecha_inicio_date.isoformat(),
            fecha_fin_date.isoformat(),
            extension='pdf',
        )
        filters = [
            f'Periodo consultado: {fecha_inicio_date} a {fecha_fin_date}',
            f'Total de registros analizados: {len(ventas)} ventas recientes',
        ]
        return self._build_document(
            'Reporte de ventas por periodo',
            filters,
            story,
            filename,
        )

    def generar_pdf_cierre_caja(self, cierre_id: int) -> GeneratedReportFile:
        """
        Genera el PDF detallado de un cierre de caja.
        """
        cierre = CierreCajaService.obtener_detalle_cierre(cierre_id)
        payment_rows = [
            ['Efectivo', self._format_currency(cierre.total_efectivo)],
            ['Tarjeta', self._format_currency(cierre.total_tarjeta)],
            [
                'Transferencia',
                self._format_currency(cierre.total_transferencia),
            ],
            ['Credito', self._format_currency(cierre.total_credito)],
            ['Abonos efectivo', self._format_currency(cierre.total_abonos)],
        ]
        gasto_rows = []
        for key, value in cierre.gastos_operativos.items():
            if key == 'total':
                continue

            if isinstance(value, dict):
                monto = value.get('monto', 0)
                descripcion = value.get('descripcion', '')
            else:
                monto = value
                descripcion = ''

            gasto_rows.append([
                key.replace('_', ' ').title(),
                self._format_currency(monto),
                descripcion or '-',
            ])

        categoria_rows = [
            [categoria, self._format_currency(total)]
            for categoria, total in sorted(
                cierre.ventas_por_categoria.items(),
                key=lambda item: item[1],
                reverse=True,
            )
        ]
        story: list[Any] = []
        story.append(self._section_title('Control de efectivo'))
        story.append(
            self._metric_table(
                [
                    ['Concepto', 'Valor', 'Referencia'],
                    [
                        'Total ventas',
                        self._format_currency(cierre.total_ventas),
                        cierre.fecha_cierre.isoformat(),
                    ],
                    [
                        'Efectivo esperado',
                        self._format_currency(cierre.efectivo_esperado),
                        'Ventas en efectivo + abonos',
                    ],
                    [
                        'Efectivo real',
                        self._format_currency(cierre.efectivo_real),
                        'Conteo registrado',
                    ],
                    [
                        'Diferencia',
                        self._format_currency(cierre.diferencia),
                        (
                            'Sobrante'
                            if cierre.diferencia >= ZERO else 'Faltante'
                        ),
                    ],
                ],
            ),
        )
        story.append(self._section_title('Ventas por metodo de pago'))
        story.append(
            self._data_table(
                ('Metodo', 'Total'),
                payment_rows,
                col_widths=[8.5 * cm, 6.0 * cm],
            ),
        )
        metodo_labels = [row[0] for row in payment_rows]
        metodo_values = [
            float(cierre.total_efectivo),
            float(cierre.total_tarjeta),
            float(cierre.total_transferencia),
            float(cierre.total_credito),
            float(cierre.total_abonos),
        ]
        metodo_chart = self._pie_chart(
            'Composicion del cierre',
            metodo_labels,
            metodo_values,
        )
        if metodo_chart is not None:
            story.append(Spacer(1, 0.2 * cm))
            story.append(metodo_chart)

        story.append(self._section_title('Gastos operativos'))
        story.append(
            self._data_table(
                ('Concepto', 'Monto', 'Descripcion'),
                gasto_rows or [['Sin gastos', '$0.00', '-']],
                col_widths=[4.6 * cm, 3.2 * cm, 6.7 * cm],
            ),
        )
        story.append(
            self._body(
                'Total de gastos operativos: '
                f'{self._format_currency(cierre.total_gastos)}',
            ),
        )
        story.append(self._section_title('Ventas por categoria'))
        story.append(
            self._data_table(
                ('Categoria', 'Total'),
                categoria_rows or [['Sin categoria', '$0.00']],
                col_widths=[8.5 * cm, 6.0 * cm],
            ),
        )
        categoria_labels = [row[0] for row in categoria_rows[:10]]
        categoria_values = [
            float(total)
            for _, total in sorted(
                cierre.ventas_por_categoria.items(),
                key=lambda item: item[1],
                reverse=True,
            )[:10]
        ]
        categoria_chart = self._bar_chart(
            'Categorias con mayor participacion',
            categoria_labels,
            categoria_values,
        )
        if categoria_chart is not None:
            story.append(Spacer(1, 0.2 * cm))
            story.append(categoria_chart)

        if cierre.observaciones:
            story.append(self._section_title('Observaciones'))
            story.append(self._body(cierre.observaciones))

        filename = self._build_filename(
            'cierre-caja',
            cierre.fecha_cierre.isoformat(),
            extension='pdf',
        )
        filters = [
            f'Fecha de cierre: {cierre.fecha_cierre}',
            f'Usuario responsable: {cierre.usuario_cierre.username}',
        ]
        return self._build_document(
            'Reporte de cierre de caja',
            filters,
            story,
            filename,
        )

    def generar_pdf_inventario_valorizado(self) -> GeneratedReportFile:
        """
        Genera un PDF de valorizacion del inventario actual.
        """
        resumen = ReporteEstadisticasService.valor_total_inventario()
        valor_compra_expr = ExpressionWrapper(
            F('precio_compra') * F('existencias'),
            output_field=DecimalField(max_digits=16, decimal_places=2),
        )
        valor_venta_expr = ExpressionWrapper(
            F('precio_venta') * F('existencias'),
            output_field=DecimalField(max_digits=16, decimal_places=2),
        )
        productos = list(
            Producto.objects.select_related('categoria').annotate(
                valor_compra_total=Coalesce(
                    valor_compra_expr,
                    ZERO,
                    output_field=DecimalField(
                        max_digits=16,
                        decimal_places=2,
                    ),
                ),
                valor_venta_total=Coalesce(
                    valor_venta_expr,
                    ZERO,
                    output_field=DecimalField(
                        max_digits=16,
                        decimal_places=2,
                    ),
                ),
            ).order_by('-valor_compra_total', 'nombre')[:20]
        )
        story: list[Any] = []
        story.append(self._section_title('Resumen de valorizacion'))
        story.append(
            self._metric_table(
                [
                    ['Indicador', 'Valor', 'Referencia'],
                    [
                        'Valor compra',
                        self._format_currency(resumen['valor_compra']),
                        'Costo del inventario actual',
                    ],
                    [
                        'Valor venta',
                        self._format_currency(resumen['valor_venta']),
                        'Valor comercial potencial',
                    ],
                    [
                        'Margen potencial',
                        self._format_currency(
                            resumen['margen_potencial'],
                        ),
                        'Valor venta - valor compra',
                    ],
                    [
                        'Existencias totales',
                        self._format_number(resumen['total_existencias']),
                        'Suma de unidades disponibles',
                    ],
                ],
            ),
        )
        story.append(self._section_title('Productos con mayor valor en stock'))
        producto_rows = [
            [
                producto.codigo_interno_formateado or '',
                producto.nombre,
                (
                    producto.categoria.nombre
                    if producto.categoria else 'Sin categoria'
                ),
                self._format_number(producto.existencias),
                self._format_currency(producto.valor_compra_total),
                self._format_currency(producto.valor_venta_total),
            ]
            for producto in productos
        ]
        story.append(
            self._data_table(
                (
                    'Codigo',
                    'Producto',
                    'Categoria',
                    'Existencias',
                    'Valor compra',
                    'Valor venta',
                ),
                producto_rows or [['', 'Sin datos', '', '0', '$0.00', '$0.00']],
                col_widths=[
                    2.2 * cm,
                    5.2 * cm,
                    3.0 * cm,
                    2.2 * cm,
                    2.9 * cm,
                    2.9 * cm,
                ],
            ),
        )
        labels = [producto.nombre[:15] for producto in productos[:10]]
        values = [
            float(self._quantize(producto.valor_compra_total))
            for producto in productos[:10]
        ]
        chart = self._bar_chart(
            'Top 10 productos por valor de compra',
            labels,
            values,
        )
        if chart is not None:
            story.append(Spacer(1, 0.2 * cm))
            story.append(chart)

        filename = self._build_filename(
            'inventario-valorizado',
            timezone.localdate().isoformat(),
            extension='pdf',
        )
        filters = [
            f'Fecha de corte: {timezone.localdate()}',
            f'Cantidad de productos: {resumen["cantidad_productos"]}',
        ]
        return self._build_document(
            'Reporte de inventario valorizado',
            filters,
            story,
            filename,
        )

    def generar_pdf_cuentas_por_cobrar(self) -> GeneratedReportFile:
        """
        Genera el PDF ejecutivo de cartera.
        """
        resumen = ReporteEstadisticasService.total_cuentas_por_cobrar()
        antiguedad = ReporteEstadisticasService.antiguedad_cartera()
        ventas = list(
            Venta.objects.select_related('cliente').filter(
                estado=Venta.Estado.TERMINADA,
                saldo_pendiente__gt=ZERO,
            ).order_by('-saldo_pendiente', 'fecha_venta')[:25]
        )
        story: list[Any] = []
        story.append(self._section_title('Resumen de cartera'))
        story.append(
            self._metric_table(
                [
                    ['Indicador', 'Valor', 'Referencia'],
                    [
                        'Total cartera',
                        self._format_currency(resumen['total_cartera']),
                        'Saldo pendiente actual',
                    ],
                    [
                        'Ventas con saldo',
                        str(resumen['cantidad_ventas']),
                        'Documentos abiertos',
                    ],
                    [
                        'Clientes con saldo',
                        str(resumen['clientes_con_saldo']),
                        'Clientes distintos',
                    ],
                    [
                        'Ticket promedio pendiente',
                        self._format_currency(
                            resumen['ticket_promedio_pendiente'],
                        ),
                        'Promedio por venta abierta',
                    ],
                ],
            ),
        )
        story.append(self._section_title('Antiguedad de cartera'))
        bucket_rows = [
            [
                item['label'],
                self._format_currency(item['total']),
                str(item['cantidad_ventas']),
            ]
            for item in antiguedad['distribucion']
        ]
        story.append(
            self._data_table(
                ('Bucket', 'Total', 'Ventas'),
                bucket_rows or [['Sin datos', '$0.00', '0']],
                col_widths=[6.0 * cm, 5.0 * cm, 4.0 * cm],
            ),
        )
        bucket_labels, bucket_values = self._chart_points(
            antiguedad['distribucion'],
            'label',
            'total',
            limit=5,
        )
        bucket_chart = self._pie_chart(
            'Distribucion de la cartera',
            bucket_labels,
            bucket_values,
        )
        if bucket_chart is not None:
            story.append(Spacer(1, 0.2 * cm))
            story.append(bucket_chart)

        story.append(self._section_title('Ventas con mayor saldo pendiente'))
        detalle_rows = [
            [
                venta.numero_venta,
                self._cliente_nombre(venta),
                self._format_date(
                    timezone.localtime(venta.fecha_venta).date(),
                ),
                self._format_currency(venta.saldo_pendiente),
                str(max(self._dias_vencidos(venta), 0)),
            ]
            for venta in ventas
        ]
        story.append(
            self._data_table(
                (
                    'Venta',
                    'Cliente',
                    'Fecha',
                    'Saldo pendiente',
                    'Dias vencidos',
                ),
                detalle_rows or [['Sin datos', '', '', '$0.00', '0']],
                col_widths=[2.6 * cm, 6.0 * cm, 2.8 * cm, 3.0 * cm, 2.6 * cm],
            ),
        )

        filename = self._build_filename(
            'cuentas-por-cobrar',
            timezone.localdate().isoformat(),
            extension='pdf',
        )
        filters = [
            f'Fecha de corte: {antiguedad["fecha_corte"]}',
            'Incluye ventas terminadas con saldo pendiente.',
        ]
        return self._build_document(
            'Reporte de cuentas por cobrar',
            filters,
            story,
            filename,
        )


class ExcelReportGenerator(BaseReportGenerator):
    """Generador especializado de reportes Excel."""

    def _new_workbook(self, title: str) -> tuple[Workbook, Any]:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = title
        return workbook, worksheet

    def _apply_title(self, worksheet, title: str, subtitle: str) -> None:
        worksheet.merge_cells('A1:H1')
        worksheet['A1'] = title
        worksheet['A1'].font = EXCEL_TITLE_FONT
        worksheet['A1'].alignment = ALIGN_LEFT
        worksheet.merge_cells('A2:H2')
        worksheet['A2'] = subtitle
        worksheet['A2'].font = EXCEL_META_FONT
        worksheet['A2'].alignment = ALIGN_LEFT

    def _apply_header(self, worksheet, row_number: int, headers: Sequence[str]):
        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=row_number, column=column_index)
            cell.value = header
            cell.font = EXCEL_HEADER_FONT
            cell.fill = EXCEL_HEADER_FILL
            cell.alignment = ALIGN_CENTER
            cell.border = EXCEL_BORDER

    def _style_body_row(self, worksheet, row_number: int, total_columns: int):
        fill = EXCEL_ALT_FILL if row_number % 2 == 0 else None

        for column_index in range(1, total_columns + 1):
            cell = worksheet.cell(row=row_number, column=column_index)
            cell.font = EXCEL_BODY_FONT
            cell.border = EXCEL_BORDER
            if fill is not None:
                cell.fill = fill

    def _set_column_widths(
        self,
        worksheet,
        widths: Sequence[float],
    ) -> None:
        for index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = width

    def _finalize_table(
        self,
        worksheet,
        header_row: int,
        freeze_cell: str,
    ) -> None:
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = freeze_cell
        worksheet.sheet_view.showGridLines = False

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.row <= 2 and cell.column > 8:
                    break

                if cell.number_format == 'General' and isinstance(
                    cell.value,
                    (int, float, Decimal),
                ):
                    cell.alignment = ALIGN_RIGHT

    def _save_workbook(
        self,
        workbook: Workbook,
        filename: str,
    ) -> GeneratedReportFile:
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return self._build_excel_result(filename, buffer)

    def _write_summary_block(
        self,
        worksheet,
        start_row: int,
        title: str,
        rows: Sequence[Sequence[Any]],
    ) -> int:
        worksheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=4,
        )
        title_cell = worksheet.cell(row=start_row, column=1)
        title_cell.value = title
        title_cell.font = EXCEL_SUBHEADER_FONT
        title_cell.fill = EXCEL_SUBHEADER_FILL
        title_cell.alignment = ALIGN_LEFT
        title_cell.border = EXCEL_BORDER

        row_number = start_row + 1
        for label, value in rows:
            worksheet.cell(row=row_number, column=1, value=label)
            worksheet.cell(row=row_number, column=2, value=value)
            worksheet.cell(row=row_number, column=1).font = EXCEL_BODY_FONT
            worksheet.cell(row=row_number, column=2).font = EXCEL_BODY_FONT
            worksheet.cell(row=row_number, column=1).border = EXCEL_BORDER
            worksheet.cell(row=row_number, column=2).border = EXCEL_BORDER
            worksheet.cell(row=row_number, column=2).alignment = ALIGN_RIGHT
            row_number += 1

        return row_number + 1

    def _insert_bar_chart(
        self,
        worksheet,
        title: str,
        data_column: int,
        categories_column: int,
        min_row: int,
        max_row: int,
        anchor: str,
    ) -> None:
        if max_row < min_row:
            return

        chart = BarChart()
        chart.title = title
        chart.height = 8
        chart.width = 14
        chart.y_axis.title = 'Valor'
        chart.x_axis.title = 'Categoria'
        data = Reference(
            worksheet,
            min_col=data_column,
            min_row=min_row - 1,
            max_row=max_row,
        )
        categories = Reference(
            worksheet,
            min_col=categories_column,
            min_row=min_row,
            max_row=max_row,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        worksheet.add_chart(chart, anchor)

    def _insert_line_chart(
        self,
        worksheet,
        title: str,
        data_column: int,
        categories_column: int,
        min_row: int,
        max_row: int,
        anchor: str,
    ) -> None:
        if max_row < min_row:
            return

        chart = LineChart()
        chart.title = title
        chart.height = 8
        chart.width = 14
        chart.y_axis.title = 'Valor'
        chart.x_axis.title = 'Fecha'
        data = Reference(
            worksheet,
            min_col=data_column,
            min_row=min_row - 1,
            max_row=max_row,
        )
        categories = Reference(
            worksheet,
            min_col=categories_column,
            min_row=min_row,
            max_row=max_row,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        worksheet.add_chart(chart, anchor)

    def _insert_pie_chart(
        self,
        worksheet,
        title: str,
        data_column: int,
        categories_column: int,
        min_row: int,
        max_row: int,
        anchor: str,
    ) -> None:
        if max_row < min_row:
            return

        chart = PieChart()
        chart.title = title
        chart.height = 8
        chart.width = 12
        data = Reference(
            worksheet,
            min_col=data_column,
            min_row=min_row - 1,
            max_row=max_row,
        )
        categories = Reference(
            worksheet,
            min_col=categories_column,
            min_row=min_row,
            max_row=max_row,
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        worksheet.add_chart(chart, anchor)

    def generar_excel_ventas_detallado(
        self,
        fecha_inicio: Any,
        fecha_fin: Any,
    ) -> GeneratedReportFile:
        """
        Genera el Excel detallado de ventas por periodo.
        """
        fecha_inicio_date = self._parse_date(fecha_inicio, 'fecha_inicio')
        fecha_fin_date = self._parse_date(fecha_fin, 'fecha_fin')
        estadisticas = ReporteEstadisticasService.estadisticas_ventas_periodo(
            fecha_inicio_date,
            fecha_fin_date,
        )
        metodos = ReporteEstadisticasService.ventas_por_metodo_pago(
            fecha_inicio_date,
            fecha_fin_date,
        )
        series = ReporteEstadisticasService.ventas_por_dia(
            fecha_inicio_date,
            fecha_fin_date,
        )['series']
        detalles = list(
            self._detalles_queryset_periodo(
                fecha_inicio_date,
                fecha_fin_date,
            ).select_related(
                'venta',
                'venta__cliente',
                'venta__usuario_registro',
                'producto',
                'producto__categoria',
            ).order_by(
                'venta__fecha_venta',
                'venta__numero_venta',
                'producto__nombre',
            )
        )
        workbook, resumen_ws = self._new_workbook('Resumen')
        detalle_ws = workbook.create_sheet('Detalle ventas')
        self._apply_title(
            resumen_ws,
            'Reporte detallado de ventas',
            (
                f'Periodo: {fecha_inicio_date} a {fecha_fin_date} | '
                f'Generado: {self._format_datetime(timezone.now())}'
            ),
        )

        resumen = estadisticas['resumen']
        next_row = self._write_summary_block(
            resumen_ws,
            4,
            'Metricas principales',
            [
                ('Total ventas', resumen['total_ventas']),
                ('Cantidad ventas', resumen['cantidad_ventas']),
                ('Ticket promedio', resumen['ticket_promedio']),
            ],
        )
        next_row = self._write_summary_block(
            resumen_ws,
            next_row,
            'Comparacion con periodo anterior',
            [
                (
                    'Variacion total ventas (%)',
                    (
                        estadisticas['comparacion_periodo_anterior'][
                            'total_ventas'
                        ]['variacion_porcentual']
                        or 0
                    ) / 100,
                ),
                (
                    'Variacion cantidad ventas (%)',
                    (
                        estadisticas['comparacion_periodo_anterior'][
                            'cantidad_ventas'
                        ]['variacion_porcentual']
                        or 0
                    ) / 100,
                ),
                (
                    'Variacion ticket promedio (%)',
                    (
                        estadisticas['comparacion_periodo_anterior'][
                            'ticket_promedio'
                        ]['variacion_porcentual']
                        or 0
                    ) / 100,
                ),
            ],
        )

        for row_number in range(5, next_row):
            value_cell = resumen_ws.cell(row=row_number, column=2)
            if row_number in (5, 7):
                value_cell.number_format = NUMBER_FORMAT_CURRENCY
            elif row_number in (10, 11, 12):
                value_cell.number_format = NUMBER_FORMAT_PERCENT

        trend_row = next_row
        self._apply_header(
            resumen_ws,
            trend_row,
            ('Fecha', 'Total ventas', 'Cantidad ventas', 'Ticket promedio'),
        )
        for index, item in enumerate(series, start=trend_row + 1):
            resumen_ws.cell(row=index, column=1, value=item['fecha'])
            resumen_ws.cell(
                row=index,
                column=2,
                value=item['total_ventas'],
            )
            resumen_ws.cell(
                row=index,
                column=3,
                value=item['cantidad_ventas'],
            )
            resumen_ws.cell(
                row=index,
                column=4,
                value=item['ticket_promedio'],
            )
            self._style_body_row(resumen_ws, index, 4)
            resumen_ws.cell(row=index, column=2).number_format = (
                NUMBER_FORMAT_CURRENCY
            )
            resumen_ws.cell(row=index, column=4).number_format = (
                NUMBER_FORMAT_CURRENCY
            )

        trend_end_row = trend_row + len(series)
        self._insert_line_chart(
            resumen_ws,
            'Tendencia diaria de ventas',
            2,
            1,
            trend_row + 1,
            trend_end_row,
            'F4',
        )

        metodo_row = trend_end_row + 3
        self._apply_header(
            resumen_ws,
            metodo_row,
            ('Metodo', 'Total vendido', 'Participacion'),
        )
        distribucion_metodos = metodos.get('distribucion', [])
        for index, item in enumerate(
            distribucion_metodos,
            start=metodo_row + 1,
        ):
            resumen_ws.cell(row=index, column=1, value=item['metodo_pago'])
            resumen_ws.cell(row=index, column=2, value=item['total_vendido'])
            resumen_ws.cell(
                row=index,
                column=3,
                value=(item['porcentaje'] or 0) / 100,
            )
            self._style_body_row(resumen_ws, index, 3)
            resumen_ws.cell(row=index, column=2).number_format = (
                NUMBER_FORMAT_CURRENCY
            )
            resumen_ws.cell(row=index, column=3).number_format = (
                NUMBER_FORMAT_PERCENT
            )

        metodo_end_row = metodo_row + len(distribucion_metodos)
        self._insert_pie_chart(
            resumen_ws,
            'Participacion por metodo de pago',
            2,
            1,
            metodo_row + 1,
            metodo_end_row,
            'F20',
        )
        self._set_column_widths(
            resumen_ws,
            (18, 18, 16, 18, 14, 14, 14, 14),
        )
        self._finalize_table(resumen_ws, trend_row, 'A5')

        detalle_headers = (
            'Venta',
            'Fecha',
            'Cliente',
            'Producto',
            'Categoria',
            'Cantidad',
            'Precio unitario',
            'Subtotal',
            'IVA',
            'Descuento',
            'Total detalle',
            'Metodo pago',
            'Estado pago',
            'Usuario',
        )
        self._apply_title(
            detalle_ws,
            'Detalle transaccional de ventas',
            (
                f'Periodo: {fecha_inicio_date} a {fecha_fin_date} | '
                'Incluye una fila por producto vendido.'
            ),
        )
        header_row = 4
        self._apply_header(detalle_ws, header_row, detalle_headers)

        for index, detalle in enumerate(detalles, start=header_row + 1):
            venta = detalle.venta
            detalle_ws.cell(row=index, column=1, value=venta.numero_venta)
            detalle_ws.cell(
                row=index,
                column=2,
                value=self._format_datetime(venta.fecha_venta),
            )
            detalle_ws.cell(
                row=index,
                column=3,
                value=self._cliente_nombre(venta),
            )
            detalle_ws.cell(row=index, column=4, value=detalle.producto.nombre)
            detalle_ws.cell(
                row=index,
                column=5,
                value=(
                    detalle.producto.categoria.nombre
                    if detalle.producto.categoria else 'Sin categoria'
                ),
            )
            detalle_ws.cell(row=index, column=6, value=float(detalle.cantidad))
            detalle_ws.cell(
                row=index,
                column=7,
                value=float(detalle.precio_unitario),
            )
            detalle_ws.cell(
                row=index,
                column=8,
                value=float(detalle.subtotal),
            )
            detalle_ws.cell(row=index, column=9, value=float(detalle.iva))
            detalle_ws.cell(
                row=index,
                column=10,
                value=float(detalle.descuento),
            )
            detalle_ws.cell(row=index, column=11, value=float(detalle.total))
            detalle_ws.cell(
                row=index,
                column=12,
                value=venta.get_metodo_pago_display(),
            )
            detalle_ws.cell(
                row=index,
                column=13,
                value=venta.get_estado_pago_display(),
            )
            detalle_ws.cell(
                row=index,
                column=14,
                value=venta.usuario_registro.username,
            )
            self._style_body_row(detalle_ws, index, len(detalle_headers))

            for numeric_column in (7, 8, 9, 10, 11):
                detalle_ws.cell(
                    row=index,
                    column=numeric_column,
                ).number_format = NUMBER_FORMAT_CURRENCY

            detalle_ws.cell(row=index, column=6).number_format = (
                NUMBER_FORMAT_QUANTITY
            )

        total_row = header_row + len(detalles) + 1
        detalle_ws.cell(row=total_row, column=1, value='Totales')
        detalle_ws.cell(
            row=total_row,
            column=6,
            value=f'=SUM(F{header_row + 1}:F{total_row - 1})',
        )
        detalle_ws.cell(
            row=total_row,
            column=8,
            value=f'=SUM(H{header_row + 1}:H{total_row - 1})',
        )
        detalle_ws.cell(
            row=total_row,
            column=9,
            value=f'=SUM(I{header_row + 1}:I{total_row - 1})',
        )
        detalle_ws.cell(
            row=total_row,
            column=10,
            value=f'=SUM(J{header_row + 1}:J{total_row - 1})',
        )
        detalle_ws.cell(
            row=total_row,
            column=11,
            value=f'=SUM(K{header_row + 1}:K{total_row - 1})',
        )
        for column in (1, 6, 8, 9, 10, 11):
            cell = detalle_ws.cell(row=total_row, column=column)
            cell.font = EXCEL_SUBHEADER_FONT
            cell.fill = EXCEL_SUBHEADER_FILL
            cell.border = EXCEL_BORDER
            if column != 1:
                cell.number_format = NUMBER_FORMAT_CURRENCY

        if detalles:
            detalle_ws.conditional_formatting.add(
                f'K{header_row + 1}:K{total_row - 1}',
                ColorScaleRule(
                    start_type='min',
                    start_color='FDE9E7',
                    mid_type='percentile',
                    mid_value=50,
                    mid_color='FFF2CC',
                    end_type='max',
                    end_color='D9EAD3',
                ),
            )
        self._set_column_widths(
            detalle_ws,
            (15, 18, 22, 30, 18, 12, 14, 14, 12, 12, 14, 14, 14, 16),
        )
        self._finalize_table(detalle_ws, header_row, 'A5')

        filename = self._build_filename(
            'ventas-detallado',
            fecha_inicio_date.isoformat(),
            fecha_fin_date.isoformat(),
            extension='xlsx',
        )
        return self._save_workbook(workbook, filename)

    def generar_excel_productos_vendidos(self) -> GeneratedReportFile:
        """
        Genera el Excel historico de productos vendidos.
        """
        fecha_inicio, fecha_fin = self._historical_sales_range()
        reporte = ReporteEstadisticasService.productos_mas_vendidos(
            fecha_inicio,
            fecha_fin,
            100,
        )
        sin_movimiento = ReporteEstadisticasService.productos_sin_movimiento(30)
        workbook, resumen_ws = self._new_workbook('Productos vendidos')
        self._apply_title(
            resumen_ws,
            'Analisis historico de productos vendidos',
            (
                f'Periodo: {fecha_inicio} a {fecha_fin} | '
                f'Generado: {self._format_datetime(timezone.now())}'
            ),
        )
        headers = (
            'Producto',
            'Codigo',
            'Categoria',
            'Cantidad vendida',
            'Total vendido',
            'Margen generado',
            'Margen %',
        )
        header_row = 4
        self._apply_header(resumen_ws, header_row, headers)
        resultados = reporte.get('resultados', [])

        for index, item in enumerate(resultados, start=header_row + 1):
            resumen_ws.cell(row=index, column=1, value=item['nombre'])
            resumen_ws.cell(
                row=index,
                column=2,
                value=item['codigo_interno'],
            )
            resumen_ws.cell(
                row=index,
                column=3,
                value=item['categoria'],
            )
            resumen_ws.cell(
                row=index,
                column=4,
                value=item['cantidad_vendida'],
            )
            resumen_ws.cell(
                row=index,
                column=5,
                value=item['total_vendido'],
            )
            resumen_ws.cell(
                row=index,
                column=6,
                value=item['margen_generado'],
            )
            resumen_ws.cell(
                row=index,
                column=7,
                value=(
                    item['margen_generado'] / item['total_vendido']
                    if item['total_vendido'] else 0
                ),
            )
            self._style_body_row(resumen_ws, index, len(headers))
            resumen_ws.cell(row=index, column=5).number_format = (
                NUMBER_FORMAT_CURRENCY
            )
            resumen_ws.cell(row=index, column=6).number_format = (
                NUMBER_FORMAT_CURRENCY
            )
            resumen_ws.cell(row=index, column=7).number_format = (
                NUMBER_FORMAT_PERCENT
            )

        total_row = header_row + len(resultados) + 1
        resumen_ws.cell(row=total_row, column=1, value='Totales')
        resumen_ws.cell(
            row=total_row,
            column=4,
            value=f'=SUM(D{header_row + 1}:D{total_row - 1})',
        )
        resumen_ws.cell(
            row=total_row,
            column=5,
            value=f'=SUM(E{header_row + 1}:E{total_row - 1})',
        )
        resumen_ws.cell(
            row=total_row,
            column=6,
            value=f'=SUM(F{header_row + 1}:F{total_row - 1})',
        )
        for column in (1, 4, 5, 6):
            cell = resumen_ws.cell(row=total_row, column=column)
            cell.font = EXCEL_SUBHEADER_FONT
            cell.fill = EXCEL_SUBHEADER_FILL
            cell.border = EXCEL_BORDER
            if column in (5, 6):
                cell.number_format = NUMBER_FORMAT_CURRENCY

        self._insert_bar_chart(
            resumen_ws,
            'Top productos por total vendido',
            5,
            1,
            header_row + 1,
            min(total_row - 1, header_row + 15),
            'I4',
        )
        if resultados:
            resumen_ws.conditional_formatting.add(
                f'G{header_row + 1}:G{total_row - 1}',
                ColorScaleRule(
                    start_type='min',
                    start_color='FDE9E7',
                    mid_type='percentile',
                    mid_value=50,
                    mid_color='FFF2CC',
                    end_type='max',
                    end_color='D9EAD3',
                ),
            )
        self._set_column_widths(
            resumen_ws,
            (30, 12, 18, 14, 16, 16, 12, 12),
        )
        self._finalize_table(resumen_ws, header_row, 'A5')

        sin_movimiento_ws = workbook.create_sheet('Sin movimiento')
        self._apply_title(
            sin_movimiento_ws,
            'Productos sin movimiento en 30 dias',
            'Ayuda operativa para depurar catalogo y promociones.',
        )
        headers_sin_mov = (
            'Producto',
            'Codigo',
            'Categoria',
            'Existencias',
            'Valor inventario',
        )
        sin_header_row = 4
        self._apply_header(sin_movimiento_ws, sin_header_row, headers_sin_mov)
        for index, item in enumerate(
            sin_movimiento.get('resultados', []),
            start=sin_header_row + 1,
        ):
            sin_movimiento_ws.cell(row=index, column=1, value=item['nombre'])
            sin_movimiento_ws.cell(
                row=index,
                column=2,
                value=item['codigo_interno'],
            )
            sin_movimiento_ws.cell(
                row=index,
                column=3,
                value=item['categoria'],
            )
            sin_movimiento_ws.cell(
                row=index,
                column=4,
                value=item['existencias'],
            )
            sin_movimiento_ws.cell(
                row=index,
                column=5,
                value=item['valor_inventario'],
            )
            self._style_body_row(sin_movimiento_ws, index, len(headers_sin_mov))
            sin_movimiento_ws.cell(row=index, column=5).number_format = (
                NUMBER_FORMAT_CURRENCY
            )

        resultados_sin_movimiento = sin_movimiento.get('resultados', [])
        if resultados_sin_movimiento:
            sin_movimiento_ws.conditional_formatting.add(
                f'D{sin_header_row + 1}:'
                f'D{sin_header_row + len(resultados_sin_movimiento)}',
                CellIsRule(
                    operator='greaterThan',
                    formula=['0'],
                    fill=EXCEL_ALERT_FILL,
                ),
            )
        self._set_column_widths(
            sin_movimiento_ws,
            (30, 12, 18, 12, 16),
        )
        self._finalize_table(sin_movimiento_ws, sin_header_row, 'A5')

        filename = self._build_filename(
            'productos-vendidos',
            fecha_inicio.isoformat(),
            fecha_fin.isoformat(),
            extension='xlsx',
        )
        return self._save_workbook(workbook, filename)

    def generar_excel_clientes_top(
        self,
        fecha_inicio: Any,
        fecha_fin: Any,
        limite: int = 100,
    ) -> GeneratedReportFile:
        """
        Genera el Excel de mejores clientes para un periodo.
        """
        fecha_inicio_date = self._parse_date(fecha_inicio, 'fecha_inicio')
        fecha_fin_date = self._parse_date(fecha_fin, 'fecha_fin')
        reporte = ReporteEstadisticasService.mejores_clientes(
            fecha_inicio_date,
            fecha_fin_date,
            limite,
        )
        workbook, worksheet = self._new_workbook('Clientes top')
        self._apply_title(
            worksheet,
            'Mejores clientes del periodo',
            (
                f'Periodo: {fecha_inicio_date} a {fecha_fin_date} | '
                f'Generado: {self._format_datetime(timezone.now())}'
            ),
        )
        headers = (
            'Cliente',
            'Documento',
            'Total comprado',
            'Cantidad compras',
            'Ticket promedio',
        )
        header_row = 4
        self._apply_header(worksheet, header_row, headers)
        resultados = reporte.get('resultados', [])

        for index, item in enumerate(resultados, start=header_row + 1):
            worksheet.cell(row=index, column=1, value=item['nombre'])
            worksheet.cell(
                row=index,
                column=2,
                value=item['numero_documento'],
            )
            worksheet.cell(
                row=index,
                column=3,
                value=item['total_comprado'],
            )
            worksheet.cell(
                row=index,
                column=4,
                value=item['cantidad_compras'],
            )
            worksheet.cell(
                row=index,
                column=5,
                value=item['ticket_promedio'],
            )
            self._style_body_row(worksheet, index, len(headers))
            worksheet.cell(row=index, column=3).number_format = (
                NUMBER_FORMAT_CURRENCY
            )
            worksheet.cell(row=index, column=5).number_format = (
                NUMBER_FORMAT_CURRENCY
            )

        total_row = header_row + len(resultados) + 1
        worksheet.cell(row=total_row, column=1, value='Totales')
        worksheet.cell(
            row=total_row,
            column=3,
            value=f'=SUM(C{header_row + 1}:C{total_row - 1})',
        )
        worksheet.cell(
            row=total_row,
            column=4,
            value=f'=SUM(D{header_row + 1}:D{total_row - 1})',
        )
        for column in (1, 3, 4):
            cell = worksheet.cell(row=total_row, column=column)
            cell.font = EXCEL_SUBHEADER_FONT
            cell.fill = EXCEL_SUBHEADER_FILL
            cell.border = EXCEL_BORDER
        worksheet.cell(row=total_row, column=3).number_format = (
            NUMBER_FORMAT_CURRENCY
        )

        self._insert_bar_chart(
            worksheet,
            'Clientes con mayor facturacion',
            3,
            1,
            header_row + 1,
            min(total_row - 1, header_row + 15),
            'G4',
        )
        self._set_column_widths(worksheet, (28, 18, 16, 16, 16, 12, 12, 12))
        self._finalize_table(worksheet, header_row, 'A5')

        filename = self._build_filename(
            'clientes-top',
            fecha_inicio_date.isoformat(),
            fecha_fin_date.isoformat(),
            extension='xlsx',
        )
        return self._save_workbook(workbook, filename)

    def generar_excel_clientes_cartera(self) -> GeneratedReportFile:
        """
        Genera el Excel operativo de clientes con cartera.
        """
        resumen = ReporteEstadisticasService.total_cuentas_por_cobrar()
        antiguedad = ReporteEstadisticasService.antiguedad_cartera()
        ventas = list(
            Venta.objects.select_related('cliente').filter(
                estado=Venta.Estado.TERMINADA,
                saldo_pendiente__gt=ZERO,
            ).order_by('-saldo_pendiente', 'fecha_venta')
        )
        workbook, resumen_ws = self._new_workbook('Resumen cartera')
        self._apply_title(
            resumen_ws,
            'Reporte de clientes con cartera',
            (
                f'Fecha de corte: {antiguedad["fecha_corte"]} | '
                f'Generado: {self._format_datetime(timezone.now())}'
            ),
        )
        next_row = self._write_summary_block(
            resumen_ws,
            4,
            'Metricas de cartera',
            [
                ('Total cartera', resumen['total_cartera']),
                ('Ventas con saldo', resumen['cantidad_ventas']),
                ('Clientes con saldo', resumen['clientes_con_saldo']),
                (
                    'Ticket promedio pendiente',
                    resumen['ticket_promedio_pendiente'],
                ),
            ],
        )
        for row_number in (5, 8):
            resumen_ws.cell(
                row=row_number,
                column=2,
            ).number_format = NUMBER_FORMAT_CURRENCY

        header_row = next_row
        self._apply_header(
            resumen_ws,
            header_row,
            ('Bucket', 'Total', 'Cantidad ventas'),
        )
        distribucion = antiguedad.get('distribucion', [])
        for index, item in enumerate(distribucion, start=header_row + 1):
            resumen_ws.cell(row=index, column=1, value=item['label'])
            resumen_ws.cell(row=index, column=2, value=item['total'])
            resumen_ws.cell(
                row=index,
                column=3,
                value=item['cantidad_ventas'],
            )
            self._style_body_row(resumen_ws, index, 3)
            resumen_ws.cell(row=index, column=2).number_format = (
                NUMBER_FORMAT_CURRENCY
            )

        end_row = header_row + len(distribucion)
        self._insert_pie_chart(
            resumen_ws,
            'Antiguedad de cartera',
            2,
            1,
            header_row + 1,
            end_row,
            'F4',
        )
        self._set_column_widths(
            resumen_ws,
            (18, 18, 16, 12, 12, 12, 12, 12),
        )
        self._finalize_table(resumen_ws, header_row, 'A5')

        detalle_ws = workbook.create_sheet('Detalle cartera')
        self._apply_title(
            detalle_ws,
            'Detalle de ventas con saldo pendiente',
            'Una fila por venta terminada con cartera abierta.',
        )
        detalle_headers = (
            'Venta',
            'Cliente',
            'Documento',
            'Fecha venta',
            'Dias plazo',
            'Dias vencidos',
            'Bucket',
            'Saldo pendiente',
            'Metodo pago',
        )
        detalle_header_row = 4
        self._apply_header(detalle_ws, detalle_header_row, detalle_headers)

        for index, venta in enumerate(ventas, start=detalle_header_row + 1):
            detalle_ws.cell(row=index, column=1, value=venta.numero_venta)
            detalle_ws.cell(
                row=index,
                column=2,
                value=self._cliente_nombre(venta),
            )
            detalle_ws.cell(
                row=index,
                column=3,
                value=(
                    venta.cliente.numero_documento
                    if venta.cliente else Cliente.CONSUMIDOR_FINAL_DOCUMENTO
                ),
            )
            detalle_ws.cell(
                row=index,
                column=4,
                value=self._format_datetime(venta.fecha_venta),
            )
            detalle_ws.cell(
                row=index,
                column=5,
                value=venta.cliente.dias_plazo if venta.cliente else 0,
            )
            detalle_ws.cell(
                row=index,
                column=6,
                value=max(self._dias_vencidos(venta), 0),
            )
            detalle_ws.cell(
                row=index,
                column=7,
                value=self._bucket_cartera(venta),
            )
            detalle_ws.cell(
                row=index,
                column=8,
                value=float(venta.saldo_pendiente),
            )
            detalle_ws.cell(
                row=index,
                column=9,
                value=venta.get_metodo_pago_display(),
            )
            self._style_body_row(detalle_ws, index, len(detalle_headers))
            detalle_ws.cell(row=index, column=8).number_format = (
                NUMBER_FORMAT_CURRENCY
            )

        total_row = detalle_header_row + len(ventas) + 1
        detalle_ws.cell(row=total_row, column=1, value='Total cartera')
        detalle_ws.cell(
            row=total_row,
            column=8,
            value=f'=SUM(H{detalle_header_row + 1}:H{total_row - 1})',
        )
        for column in (1, 8):
            cell = detalle_ws.cell(row=total_row, column=column)
            cell.font = EXCEL_SUBHEADER_FONT
            cell.fill = EXCEL_SUBHEADER_FILL
            cell.border = EXCEL_BORDER
        detalle_ws.cell(row=total_row, column=8).number_format = (
            NUMBER_FORMAT_CURRENCY
        )
        if ventas:
            detalle_ws.conditional_formatting.add(
                f'F{detalle_header_row + 1}:F{total_row - 1}',
                CellIsRule(
                    operator='greaterThan',
                    formula=['0'],
                    fill=EXCEL_ALERT_FILL,
                ),
            )
            detalle_ws.conditional_formatting.add(
                f'H{detalle_header_row + 1}:H{total_row - 1}',
                ColorScaleRule(
                    start_type='min',
                    start_color='D9EAD3',
                    mid_type='percentile',
                    mid_value=50,
                    mid_color='FFF2CC',
                    end_type='max',
                    end_color='FDE9E7',
                ),
            )
        self._set_column_widths(
            detalle_ws,
            (15, 28, 16, 18, 12, 12, 12, 16, 16),
        )
        self._finalize_table(detalle_ws, detalle_header_row, 'A5')

        filename = self._build_filename(
            'clientes-cartera',
            antiguedad['fecha_corte'],
            extension='xlsx',
        )
        return self._save_workbook(workbook, filename)


def generar_pdf_ventas_periodo(
    datos: Optional[dict[str, Any]],
    fecha_inicio: Any,
    fecha_fin: Any,
) -> GeneratedReportFile:
    """Fachada funcional para reporte PDF de ventas por periodo."""
    return PDFReportGenerator().generar_pdf_ventas_periodo(
        datos,
        fecha_inicio,
        fecha_fin,
    )


def generar_pdf_cierre_caja(cierre_id: int) -> GeneratedReportFile:
    """Fachada funcional para PDF de cierre de caja."""
    return PDFReportGenerator().generar_pdf_cierre_caja(cierre_id)


def generar_pdf_inventario_valorizado() -> GeneratedReportFile:
    """Fachada funcional para PDF de inventario valorizado."""
    return PDFReportGenerator().generar_pdf_inventario_valorizado()


def generar_pdf_cuentas_por_cobrar() -> GeneratedReportFile:
    """Fachada funcional para PDF de cuentas por cobrar."""
    return PDFReportGenerator().generar_pdf_cuentas_por_cobrar()


def generar_excel_ventas_detallado(
    fecha_inicio: Any,
    fecha_fin: Any,
) -> GeneratedReportFile:
    """Fachada funcional para Excel detallado de ventas."""
    return ExcelReportGenerator().generar_excel_ventas_detallado(
        fecha_inicio,
        fecha_fin,
    )


def generar_excel_productos_vendidos(
    fecha_inicio: Any = None,
    fecha_fin: Any = None,
) -> GeneratedReportFile:
    """Fachada funcional para Excel de productos vendidos."""
    generator = ExcelReportGenerator()
    if fecha_inicio is None or fecha_fin is None:
        return generator.generar_excel_productos_vendidos()

    fecha_inicio_date = generator._parse_date(fecha_inicio, 'fecha_inicio')
    fecha_fin_date = generator._parse_date(fecha_fin, 'fecha_fin')
    reporte = ReporteEstadisticasService.productos_mas_vendidos(
        fecha_inicio_date,
        fecha_fin_date,
        100,
    )
    workbook, worksheet = generator._new_workbook('Productos vendidos')
    generator._apply_title(
        worksheet,
        'Productos mas vendidos del periodo',
        (
            f'Periodo: {fecha_inicio_date} a {fecha_fin_date} | '
            f'Generado: {generator._format_datetime(timezone.now())}'
        ),
    )
    headers = (
        'Producto',
        'Codigo',
        'Categoria',
        'Cantidad vendida',
        'Total vendido',
        'Margen generado',
    )
    header_row = 4
    generator._apply_header(worksheet, header_row, headers)
    resultados = reporte.get('resultados', [])

    for index, item in enumerate(resultados, start=header_row + 1):
        worksheet.cell(row=index, column=1, value=item['nombre'])
        worksheet.cell(row=index, column=2, value=item['codigo_interno'])
        worksheet.cell(row=index, column=3, value=item['categoria'])
        worksheet.cell(
            row=index,
            column=4,
            value=item['cantidad_vendida'],
        )
        worksheet.cell(row=index, column=5, value=item['total_vendido'])
        worksheet.cell(
            row=index,
            column=6,
            value=item['margen_generado'],
        )
        generator._style_body_row(worksheet, index, len(headers))
        worksheet.cell(row=index, column=5).number_format = (
            NUMBER_FORMAT_CURRENCY
        )
        worksheet.cell(row=index, column=6).number_format = (
            NUMBER_FORMAT_CURRENCY
        )

    total_row = header_row + len(resultados) + 1
    worksheet.cell(row=total_row, column=1, value='Totales')
    worksheet.cell(
        row=total_row,
        column=4,
        value=f'=SUM(D{header_row + 1}:D{total_row - 1})',
    )
    worksheet.cell(
        row=total_row,
        column=5,
        value=f'=SUM(E{header_row + 1}:E{total_row - 1})',
    )
    worksheet.cell(
        row=total_row,
        column=6,
        value=f'=SUM(F{header_row + 1}:F{total_row - 1})',
    )
    for column in (1, 4, 5, 6):
        cell = worksheet.cell(row=total_row, column=column)
        cell.font = EXCEL_SUBHEADER_FONT
        cell.fill = EXCEL_SUBHEADER_FILL
        cell.border = EXCEL_BORDER
    worksheet.cell(row=total_row, column=5).number_format = (
        NUMBER_FORMAT_CURRENCY
    )
    worksheet.cell(row=total_row, column=6).number_format = (
        NUMBER_FORMAT_CURRENCY
    )

    generator._insert_bar_chart(
        worksheet,
        'Productos con mayor ingreso',
        5,
        1,
        header_row + 1,
        min(total_row - 1, header_row + 15),
        'H4',
    )
    generator._set_column_widths(worksheet, (30, 14, 18, 16, 16, 16, 12, 12))
    generator._finalize_table(worksheet, header_row, 'A5')
    filename = generator._build_filename(
        'productos-vendidos',
        fecha_inicio_date.isoformat(),
        fecha_fin_date.isoformat(),
        extension='xlsx',
    )
    return generator._save_workbook(workbook, filename)


def generar_excel_clientes_top(
    fecha_inicio: Any,
    fecha_fin: Any,
    limite: int = 100,
) -> GeneratedReportFile:
    """Fachada funcional para Excel de mejores clientes."""
    return ExcelReportGenerator().generar_excel_clientes_top(
        fecha_inicio,
        fecha_fin,
        limite,
    )


def generar_excel_clientes_cartera() -> GeneratedReportFile:
    """Fachada funcional para Excel de clientes con cartera."""
    return ExcelReportGenerator().generar_excel_clientes_cartera()
