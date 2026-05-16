from io import BytesIO
from typing import Optional

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from .models import Producto


HOJA_INVENTARIO = 'Inventario'
HOJA_INSTRUCCIONES = 'Instrucciones'

COLUMNAS_EXCEL = [
    ('N°', 8),
    ('Código Interno', 16),
    ('Código de Barras', 20),
    ('Nombre', 40),
    ('Categoría', 20),
    ('Marca', 20),
    ('Descripción', 40),
    ('Existencias', 14),
    ('Invima', 18),
    ('Precio Compra', 16),
    ('Precio Venta', 16),
    ('IVA (%)', 10),
    ('Fecha Ingreso', 18),
    ('Fecha Caducidad', 18),
]

COLUMNAS_IMPORTACION_EXCEL = [
    ('Código Interno', 16),
    ('Código de Barras', 20),
    ('Nombre', 40),
    ('Categoría', 20),
    ('Marca', 20),
    ('Descripción', 40),
    ('Existencias', 14),
    ('Invima', 18),
    ('Precio Compra', 16),
    ('Precio Venta', 16),
    ('IVA (%)', 10),
    ('Fecha Ingreso', 18),
    ('Fecha Caducidad', 18),
]

HEADER_FONT = Font(
    name='Calibri',
    size=11,
    bold=True,
    color='FFFFFF',
)
HEADER_FILL = PatternFill(
    start_color='2F5496',
    end_color='2F5496',
    fill_type='solid',
)
HEADER_ALIGNMENT = Alignment(
    horizontal='center',
    vertical='center',
    wrap_text=True,
)

CELL_FONT = Font(name='Calibri', size=10)
CELL_ALIGNMENT_TEXT = Alignment(horizontal='left', vertical='center')
CELL_ALIGNMENT_NUMBER = Alignment(horizontal='right', vertical='center')
CELL_ALIGNMENT_CENTER = Alignment(horizontal='center', vertical='center')

THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)

ALT_FILL = PatternFill(
    start_color='D6E4F0',
    end_color='D6E4F0',
    fill_type='solid',
)

FORMATO_NUMERO = '#,##0.00'
FORMATO_ENTERO = '#,##0'
FORMATO_FECHA = 'YYYY-MM-DD'


def generar_excel_inventario(
    productos: Optional[list] = None,
) -> BytesIO:
    if productos is None:
        productos = Producto.objects.select_related(
            'categoria',
        ).order_by('codigo_interno').all()

    wb = Workbook()
    ws = wb.active
    ws.title = HOJA_INVENTARIO
    ws.sheet_properties.pageSetUpPr = None
    ws.sheet_view.showGridLines = False
    _aplicar_encabezados(ws, COLUMNAS_EXCEL)

    for row_idx, producto in enumerate(productos, start=2):
        fila = _crear_fila_producto(producto, row_idx)
        es_par = row_idx % 2 == 0

        for col_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = CELL_FONT
            cell.border = THIN_BORDER

            if col_idx == 1:
                cell.alignment = CELL_ALIGNMENT_CENTER
            elif col_idx == 8:
                cell.alignment = CELL_ALIGNMENT_CENTER
                cell.number_format = FORMATO_ENTERO
            elif col_idx == 9:
                cell.alignment = CELL_ALIGNMENT_CENTER
            elif col_idx in (10, 11):
                cell.alignment = CELL_ALIGNMENT_NUMBER
                cell.number_format = FORMATO_NUMERO
            elif col_idx == 12:
                cell.alignment = CELL_ALIGNMENT_CENTER
                cell.number_format = '0.00'
            elif col_idx in (13, 14):
                cell.alignment = CELL_ALIGNMENT_CENTER
                cell.number_format = FORMATO_FECHA
            else:
                cell.alignment = CELL_ALIGNMENT_TEXT

            if es_par:
                cell.fill = ALT_FILL

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generar_plantilla_excel_inventario() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = HOJA_INVENTARIO
    ws.sheet_view.showGridLines = False
    _aplicar_encabezados(ws, COLUMNAS_IMPORTACION_EXCEL)
    ws.freeze_panes = 'A2'

    instrucciones = wb.create_sheet(title=HOJA_INSTRUCCIONES)
    instrucciones.sheet_view.showGridLines = False
    instrucciones.column_dimensions['A'].width = 110
    instrucciones['A1'] = 'Instrucciones para importar productos'
    instrucciones['A1'].font = Font(
        name='Calibri',
        size=13,
        bold=True,
        color='2F5496',
    )
    instrucciones['A3'] = (
        '1. Complete una fila por producto nuevo.\n'
        '2. Código Interno es opcional. Si coincide con un producto '
        'existente, Mallor actualizará ese registro.\n'
        '3. Nombre, Categoría, Existencias, Precio Compra y '
        'Precio Venta son obligatorios.\n'
        '4. Si la categoría no existe, Mallor la creará '
        'automáticamente dentro de la empresa activa.\n'
        '5. IVA (%) es opcional. Si queda vacío, se usará 0.\n'
        '6. Fecha Ingreso es opcional. Si queda vacía, se usará la '
        'fecha y hora actual.\n'
        '7. Fecha Caducidad puede quedar vacía.\n'
        '8. No modifique los encabezados de la hoja Inventario.\n'
        '9. El archivo debe guardarse en formato .xlsx.'
    )
    instrucciones['A3'].alignment = Alignment(
        vertical='top',
        wrap_text=True,
    )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _crear_fila_producto(
    producto: Producto,
    numero: int,
) -> list:
    fecha_ingreso = producto.fecha_ingreso
    if fecha_ingreso:
        fecha_ingreso = fecha_ingreso.date()

    return [
        numero,
        producto.codigo_interno_formateado,
        producto.codigo_barras,
        producto.nombre,
        producto.categoria.nombre if producto.categoria else '',
        producto.marca,
        producto.descripcion or '',
        int(producto.existencias) if producto.existencias else 0,
        producto.invima or '',
        float(producto.precio_compra) if producto.precio_compra else 0.0,
        float(producto.precio_venta) if producto.precio_venta else 0.0,
        float(producto.iva) if producto.iva else 0.0,
        fecha_ingreso,
        producto.fecha_caducidad,
    ]


def _aplicar_encabezados(ws, columnas) -> None:
    for col_idx, (header, width) in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30


def generar_respuesta_excel(output: BytesIO, filename: str) -> HttpResponse:
    response = HttpResponse(
        output.getvalue(),
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        ),
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename}"'
    )
    return response
