from io import BytesIO

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook, load_workbook

from inventario.models import Categoria, Producto
from tests.factories import ProductoFactory


DEFAULT_HEADERS = [
    'Código Interno',
    'Código de Barras',
    'Nombre',
    'Categoría',
    'Marca',
    'Descripción',
    'Existencias',
    'Invima',
    'Precio Compra',
    'Precio Venta',
    'IVA (%)',
    'Fecha Ingreso',
    'Fecha Caducidad',
]


def build_excel(rows, headers=None, sheet_name='Inventario'):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers or DEFAULT_HEADERS)
    for row in rows:
        worksheet.append(row)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.mark.django_db
def test_descargar_plantilla_excel_retorna_hojas_y_encabezados(
    api_client_empresa,
):
    response = api_client_empresa.get(
        '/api/inventario/productos/plantilla-excel/'
    )

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ['Inventario', 'Instrucciones']
    worksheet = workbook['Inventario']
    headers = [cell.value for cell in worksheet[1]]
    assert headers == DEFAULT_HEADERS
    assert worksheet.max_row == 1


@pytest.mark.django_db
@pytest.mark.multitenant
def test_importar_excel_crea_productos_y_categoria_en_empresa_activa(
    api_client_empresa,
    empresa_a,
    empresa_b,
):
    ProductoFactory(
        empresa=empresa_b,
        codigo_barras='7709999999999',
    )
    file_bytes = build_excel([
        [
            '',
            '7701234567890',
            'Gel Antibacterial',
            'Aseo',
            'Mallor',
            'Frasco 250ml',
            12,
            'INV-001',
            1000,
            1400,
            19,
            '2026-05-16',
            '2027-05-16',
        ],
    ])
    upload = SimpleUploadedFile(
        'productos.xlsx',
        file_bytes,
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        ),
    )

    response = api_client_empresa.post(
        '/api/inventario/productos/importar-excel/',
        {'archivo': upload},
        format='multipart',
    )

    assert response.status_code == 200
    assert response.json() == {
        'success': True,
        'message': 'Importación completada exitosamente.',
        'imported_count': 1,
    }

    categoria = Categoria.objects.get(empresa=empresa_a, nombre='ASEO')
    producto = Producto.objects.get(
        empresa=empresa_a,
        codigo_barras='7701234567890',
    )
    assert producto.nombre == 'Gel Antibacterial'
    assert producto.categoria == categoria
    assert producto.codigo_interno is not None
    assert producto.precio_compra == 1000
    assert producto.precio_venta == 1400
    assert producto.iva == 19


@pytest.mark.django_db
@pytest.mark.multitenant
def test_importar_excel_actualiza_producto_por_codigo_interno(
    api_client_empresa,
    empresa_a,
):
    producto = ProductoFactory(
        empresa=empresa_a,
        codigo_interno=23,
        codigo_barras='7700000000023',
        nombre='Arroz',
        precio_compra=2500,
        precio_venta=2500,
        existencias=10,
    )
    file_bytes = build_excel(
        [
            [
                1,
                '00000023',
                '7700000000023',
                'Arroz',
                'Granos',
                'Mallor',
                'Actualizado desde Excel',
                25,
                '',
                2600,
                3000,
                19,
                '2026-05-16',
                '',
            ],
        ],
        headers=[
            'N°',
            'Código Interno',
            'Código de Barras',
            'Nombre',
            'Categoría',
            'Marca',
            'Descripción',
            'Existencias',
            'Invima',
            'Precio Compra',
            'Precio Venta',
            'IVA (%)',
            'Fecha Ingreso',
            'Fecha Caducidad',
        ],
    )
    upload = SimpleUploadedFile(
        'productos.xlsx',
        file_bytes,
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        ),
    )

    response = api_client_empresa.post(
        '/api/inventario/productos/importar-excel/',
        {'archivo': upload},
        format='multipart',
    )

    assert response.status_code == 200
    producto.refresh_from_db()
    assert Producto.objects.filter(empresa=empresa_a).count() == 1
    assert producto.nombre == 'Arroz'
    assert producto.codigo_interno == 23
    assert str(producto.precio_compra) == '2600.00'
    assert str(producto.precio_venta) == '3000.00'
    assert str(producto.existencias) == '25.00'
    assert producto.descripcion == 'Actualizado desde Excel'


@pytest.mark.django_db
def test_importar_excel_reporta_errores_y_no_importa_parcialmente(
    api_client_empresa,
    empresa_a,
):
    ProductoFactory(
        empresa=empresa_a,
        codigo_barras='7701231231231',
    )
    file_bytes = build_excel([
        [
            '',
            '7701231231231',
            '',
            'Medicamentos',
            'Mallor',
            'Descripcion',
            'diez',
            '',
            500,
            400,
            19,
            '',
            '',
        ],
    ])
    upload = SimpleUploadedFile(
        'productos.xlsx',
        file_bytes,
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        ),
    )

    response = api_client_empresa.post(
        '/api/inventario/productos/importar-excel/',
        {'archivo': upload},
        format='multipart',
    )

    assert response.status_code == 400
    payload = response.json()
    assert payload['success'] is False
    assert 'No se pudo importar el archivo' in payload['message']
    assert payload['errors'] == [
        {
            'row': 2,
            'column': 'Nombre',
            'value': '',
            'error': 'Este campo es obligatorio.',
        },
        {
            'row': 2,
            'column': 'Existencias',
            'value': 'diez',
            'error': 'Debe ser un número entero mayor o igual a 0.',
        },
        {
            'row': 2,
            'column': 'Precio Venta',
            'value': '400',
            'error': 'No puede ser menor que Precio Compra.',
        },
        {
            'row': 2,
            'column': 'Código de Barras',
            'value': '7701231231231',
            'error': 'Ya existe un producto con este código de barras.',
        },
    ]
    assert Producto.objects.filter(empresa=empresa_a).count() == 1
