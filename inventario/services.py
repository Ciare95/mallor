from typing import List, Dict, Optional, Any
from decimal import Decimal
from datetime import date, datetime
from django.db import transaction
from django.db.models import Q, Sum, DecimalField, F
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from openpyxl import load_workbook

from core.exceptions import (
    InventarioError,
    ProductoNoEncontradoError,
    ProductoDuplicadoError,
    ProductoConMovimientosError,
    StockInsuficienteError,
    FacturaNoEncontradaError,
    FacturaYaProcesadaError,
    FacturaSinDetallesError,
    CategoriaNoEncontradaError,
    CategoriaConProductosError,
    ImportacionInventarioError,
)
from empresa.context import get_empresa_actual_or_default

from .models import (
    Categoria,
    Producto,
    FacturaCompra,
    DetalleFacturaCompra,
    HistorialInventario,
)
from .utils import COLUMNAS_IMPORTACION_EXCEL, HOJA_INVENTARIO
from proveedor.models import Proveedor
from usuario.models import Usuario

DEFAULT_SALE_PRICING_RULES = {
    'umbral': Decimal('1000'),
    'margen_menor_igual': Decimal('119'),
    'margen_mayor': Decimal('69'),
}


def normalize_sale_pricing_rules(
    pricing_rules: Optional[Dict[str, Any]] = None
) -> Dict[str, Decimal]:
    rule = pricing_rules or {}

    def parse_decimal(key: str, default: Decimal) -> Decimal:
        value = rule.get(key, default)
        try:
            return Decimal(str(value))
        except (ValueError, TypeError, ArithmeticError):
            return default

    return {
        'umbral': parse_decimal(
            'umbral',
            DEFAULT_SALE_PRICING_RULES['umbral'],
        ),
        'margen_menor_igual': parse_decimal(
            'margen_menor_igual',
            DEFAULT_SALE_PRICING_RULES['margen_menor_igual'],
        ),
        'margen_mayor': parse_decimal(
            'margen_mayor',
            DEFAULT_SALE_PRICING_RULES['margen_mayor'],
        ),
    }


def calculate_suggested_sale_price(
    base_price: Decimal,
    pricing_rules: Optional[Dict[str, Any]] = None,
) -> Decimal:
    q = Decimal('0.01')
    if base_price <= 0:
        return Decimal('0.00')

    rule = normalize_sale_pricing_rules(pricing_rules)
    markup = (
        rule['margen_menor_igual']
        if base_price <= rule['umbral']
        else rule['margen_mayor']
    )
    return (
        base_price * (Decimal('1') + (markup / Decimal('100')))
    ).quantize(q)


class CategoriaService:
    """
    Servicio para gestionar la lógica de negocio de categorías.
    """

    @staticmethod
    def crear_categoria(data: Dict[str, Any]) -> Categoria:
        empresa = get_empresa_actual_or_default()
        nombre = data.get('nombre', '').strip().upper()
        if Categoria.objects.filter(empresa=empresa, nombre__iexact=nombre).exists():
            raise ProductoDuplicadoError('nombre', nombre)
        return Categoria.objects.create(
            empresa=empresa,
            nombre=nombre,
            descripcion=data.get('descripcion', ''),
        )

    @staticmethod
    def obtener_categoria(categoria_id: int) -> Categoria:
        try:
            return Categoria.objects.get(
                id=categoria_id,
                empresa=get_empresa_actual_or_default(),
            )
        except Categoria.DoesNotExist:
            raise CategoriaNoEncontradaError(categoria_id)

    @staticmethod
    def listar_categorias(filtros: Optional[Dict[str, Any]] = None) -> List[Categoria]:
        queryset = Categoria.objects.filter(empresa=get_empresa_actual_or_default())
        if filtros:
            q_objects = Q()
            if filtros.get('q'):
                q_objects &= Q(nombre__icontains=filtros['q'])
            if q_objects:
                queryset = queryset.filter(q_objects)
        return list(queryset.order_by('nombre'))

    @staticmethod
    @transaction.atomic
    def actualizar_categoria(categoria_id: int, data: Dict[str, Any]) -> Categoria:
        categoria = CategoriaService.obtener_categoria(categoria_id)
        nombre = data.get('nombre')
        if nombre:
            nombre = nombre.strip().upper()
            if Categoria.objects.filter(
                empresa=get_empresa_actual_or_default(),
                nombre__iexact=nombre,
            ).exclude(id=categoria_id).exists():
                raise ProductoDuplicadoError('nombre', nombre)
            categoria.nombre = nombre
        if 'descripcion' in data:
            categoria.descripcion = data['descripcion']
        categoria.save()
        return categoria

    @staticmethod
    @transaction.atomic
    def eliminar_categoria(categoria_id: int) -> None:
        categoria = CategoriaService.obtener_categoria(categoria_id)
        if categoria.producto_set.exists():
            raise CategoriaConProductosError(categoria.nombre)
        categoria.delete()


class ProductoService:
    """
    Servicio para gestionar la lógica de negocio de productos.

    Encapsula todas las reglas de negocio relacionadas con la creación,
    actualización, eliminación y consulta de productos del inventario.
    Sigue el principio de Single Responsibility (SOLID).
    """

    @staticmethod
    @transaction.atomic
    def crear_producto(
        data: Dict[str, Any],
        usuario: Optional[Usuario] = None
    ) -> Producto:
        """
        Crea un nuevo producto con validaciones de negocio.

        Args:
            data: Datos del producto a crear
            usuario: Usuario que realiza la creación (opcional)

        Returns:
            Producto: Instancia del producto creado

        Raises:
            ProductoDuplicadoError: Si el código interno o de barras ya existe
        """
        codigo_interno = data.get('codigo_interno')
        codigo_barras = data.get('codigo_barras', '')
        empresa = get_empresa_actual_or_default()
        data['empresa'] = empresa

        if codigo_interno is not None:
            if Producto.objects.filter(
                empresa=empresa,
                codigo_interno=codigo_interno,
            ).exists():
                raise ProductoDuplicadoError(
                    'código interno', codigo_interno
                )

        if codigo_barras:
            codigo_barras = codigo_barras.strip()
            data['codigo_barras'] = codigo_barras
            if Producto.objects.filter(
                empresa=empresa,
                codigo_barras=codigo_barras,
            ).exists():
                raise ProductoDuplicadoError(
                    'código de barras', codigo_barras
                )

        producto = Producto.objects.create(**data)
        return producto

    @staticmethod
    def obtener_producto(producto_id: int) -> Producto:
        """
        Obtiene un producto por su ID incluyendo relaciones.

        Args:
            producto_id: ID del producto

        Returns:
            Producto: Instancia del producto con relaciones precargadas

        Raises:
            ProductoNoEncontradoError: Si el producto no existe
        """
        try:
            return Producto.objects.select_related(
                'categoria'
            ).get(id=producto_id, empresa=get_empresa_actual_or_default())
        except Producto.DoesNotExist:
            raise ProductoNoEncontradoError(producto_id)

    @staticmethod
    def listar_productos(
        filtros: Optional[Dict[str, Any]] = None
    ) -> List[Producto]:
        """
        Lista productos aplicando filtros opcionales.

        Args:
            filtros: Diccionario con filtros (categoria, marca,
                     q, fecha_caducidad_desde, fecha_caducidad_hasta,
                     stock_min, stock_max, ordering)

        Returns:
            List[Producto]: Lista de productos filtrados
        """
        queryset = Producto.objects.select_related('categoria').filter(
            empresa=get_empresa_actual_or_default(),
        )

        if not filtros:
            return list(queryset.order_by('nombre'))

        q_objects = Q()

        if filtros.get('q'):
            q = filtros['q']
            q_objects &= (
                Q(nombre__icontains=q) |
                Q(codigo_barras__icontains=q) |
                Q(marca__icontains=q) |
                Q(invima__icontains=q)
            )
            try:
                codigo_int = int(q)
                q_objects |= Q(codigo_interno=codigo_int)
            except ValueError:
                pass

        if filtros.get('categoria_id'):
            q_objects &= Q(categoria_id=filtros['categoria_id'])

        if filtros.get('marca'):
            q_objects &= Q(marca__icontains=filtros['marca'])

        if filtros.get('fecha_caducidad_desde'):
            q_objects &= Q(
                fecha_caducidad__gte=filtros['fecha_caducidad_desde']
            )

        if filtros.get('fecha_caducidad_hasta'):
            q_objects &= Q(
                fecha_caducidad__lte=filtros['fecha_caducidad_hasta']
            )

        if filtros.get('stock_min') is not None:
            q_objects &= Q(existencias__gte=filtros['stock_min'])

        if filtros.get('stock_max') is not None:
            q_objects &= Q(existencias__lte=filtros['stock_max'])

        stock_bajo = filtros.get('stock_bajo')
        if stock_bajo in (True, 'true', 'True', '1', 1):
            q_objects &= Q(existencias__lte=F('stock_minimo'))

        if q_objects:
            queryset = queryset.filter(q_objects)

        ordering = filtros.get('ordering', 'nombre')
        ordenes_permitidos = [
            'nombre', '-nombre', 'codigo_interno', '-codigo_interno',
            'precio_compra', '-precio_compra', 'precio_venta', '-precio_venta',
            'existencias', '-existencias', 'stock_minimo', '-stock_minimo',
            'created_at', '-created_at',
        ]
        if ordering in ordenes_permitidos:
            queryset = queryset.order_by(ordering)
        else:
            queryset = queryset.order_by('nombre')

        return list(queryset)

    @staticmethod
    @transaction.atomic
    def actualizar_producto(
        producto_id: int,
        data: Dict[str, Any]
    ) -> Producto:
        """
        Actualiza un producto existente.

        Args:
            producto_id: ID del producto a actualizar
            data: Datos a actualizar

        Returns:
            Producto: Instancia del producto actualizado

        Raises:
            ProductoNoEncontradoError: Si el producto no existe
            ProductoDuplicadoError: Si el código de barras ya está en uso
        """
        producto = ProductoService.obtener_producto(producto_id)

        codigo_barras = data.get('codigo_barras')
        if codigo_barras:
            codigo_barras = codigo_barras.strip()
            data['codigo_barras'] = codigo_barras
            if Producto.objects.filter(
                empresa=get_empresa_actual_or_default(),
                codigo_barras=codigo_barras
            ).exclude(id=producto_id).exists():
                raise ProductoDuplicadoError(
                    'código de barras', codigo_barras
                )

        codigo_interno = data.get('codigo_interno')
        if codigo_interno is not None:
            if Producto.objects.filter(
                empresa=get_empresa_actual_or_default(),
                codigo_interno=codigo_interno
            ).exclude(id=producto_id).exists():
                raise ProductoDuplicadoError(
                    'código interno', codigo_interno
                )

        for campo, valor in data.items():
            if hasattr(producto, campo):
                setattr(producto, campo, valor)

        producto.save()
        return producto

    @staticmethod
    @transaction.atomic
    def eliminar_producto(producto_id: int) -> None:
        """
        Elimina un producto (solo si no tiene movimientos registrados).

        Args:
            producto_id: ID del producto a eliminar

        Raises:
            ProductoNoEncontradoError: Si el producto no existe
            ProductoConMovimientosError: Si el producto tiene movimientos
        """
        producto = ProductoService.obtener_producto(producto_id)

        if producto.historial.exists():
            raise ProductoConMovimientosError(producto_id)

        producto.delete()

    @staticmethod
    def buscar_producto(query: str) -> List[Producto]:
        """
        Búsqueda avanzada de productos por múltiples criterios.

        Args:
            query: Término de búsqueda

        Returns:
            List[Producto]: Lista de productos que coinciden
        """
        if not query or not query.strip():
            return list(
                Producto.objects.select_related('categoria').all()[:50]
            )

        q = query.strip()
        q_objects = (
            Q(nombre__icontains=q) |
            Q(codigo_barras__icontains=q) |
            Q(marca__icontains=q) |
            Q(invima__icontains=q) |
            Q(descripcion__icontains=q)
        )

        try:
            codigo_int = int(q)
            q_objects |= Q(codigo_interno=codigo_int)
        except ValueError:
            pass

        return list(
            Producto.objects.select_related('categoria')
            .filter(q_objects, empresa=get_empresa_actual_or_default())
            .order_by('nombre')[:50]
        )


class ProductoImportService:
    MAX_FILE_SIZE = 5 * 1024 * 1024
    OPTIONAL_HEADERS = {
        'Código Interno',
        'N°',
    }
    REQUIRED_HEADERS = {
        'Nombre',
        'Categoría',
        'Existencias',
        'Precio Compra',
        'Precio Venta',
    }
    EXPECTED_HEADERS = [header for header, _ in COLUMNAS_IMPORTACION_EXCEL]

    @classmethod
    def importar_desde_excel(
        cls,
        archivo,
        usuario: Optional[Usuario] = None,
    ) -> int:
        cls._validar_archivo(archivo)
        archivo.seek(0)

        try:
            workbook = load_workbook(
                filename=archivo,
                data_only=True,
            )
        except Exception as exc:
            raise ImportacionInventarioError([
                cls._build_error(
                    row=1,
                    column='Archivo',
                    value=getattr(archivo, 'name', ''),
                    error='No fue posible leer el archivo Excel.',
                ),
            ]) from exc

        worksheet = cls._obtener_hoja_trabajo(workbook)
        rows = list(worksheet.iter_rows(values_only=True))
        errors, rows_data = cls._validar_contenido(rows)
        if errors:
            raise ImportacionInventarioError(errors)

        with transaction.atomic():
            cls._crear_productos(rows_data, usuario=usuario)

        return len(rows_data)

    @classmethod
    def _validar_archivo(cls, archivo) -> None:
        nombre = (getattr(archivo, 'name', '') or '').lower()
        if not nombre.endswith('.xlsx'):
            raise ImportacionInventarioError([
                cls._build_error(
                    row=1,
                    column='Archivo',
                    value=getattr(archivo, 'name', ''),
                    error='El archivo debe estar en formato .xlsx.',
                ),
            ])

        size = getattr(archivo, 'size', 0) or 0
        if size > cls.MAX_FILE_SIZE:
            raise ImportacionInventarioError([
                cls._build_error(
                    row=1,
                    column='Archivo',
                    value=getattr(archivo, 'name', ''),
                    error=(
                        'El archivo supera el tamaño máximo permitido '
                        'de 5 MB.'
                    ),
                ),
            ])

    @classmethod
    def _obtener_hoja_trabajo(cls, workbook):
        if HOJA_INVENTARIO in workbook.sheetnames:
            return workbook[HOJA_INVENTARIO]
        return workbook[workbook.sheetnames[0]]

    @classmethod
    def _validar_contenido(cls, rows):
        if not rows:
            return [
                cls._build_error(
                    row=1,
                    column='Archivo',
                    value='',
                    error='El archivo está vacío.',
                ),
            ], []

        headers = cls._normalize_headers(rows[0])
        header_errors = cls._validar_headers(headers)
        if header_errors:
            return header_errors, []

        company = get_empresa_actual_or_default()
        existing_products_by_internal_code = {
            producto.codigo_interno: producto
            for producto in Producto.objects.filter(empresa=company)
            if producto.codigo_interno is not None
        }
        seen_barcodes = set()
        seen_internal_codes = set()
        rows_data = []
        errors = []

        for excel_row, row_values in enumerate(rows[1:], start=2):
            if cls._row_is_empty(row_values):
                continue

            row_map = {
                header: row_values[index]
                if index < len(row_values) else None
                for index, header in enumerate(headers)
            }
            row_errors, normalized = cls._validar_fila(
                excel_row=excel_row,
                row_data=row_map,
                existing_products_by_internal_code=(
                    existing_products_by_internal_code
                ),
                seen_barcodes=seen_barcodes,
                seen_internal_codes=seen_internal_codes,
            )
            errors.extend(row_errors)
            if normalized:
                rows_data.append(normalized)

        if not rows_data and not errors:
            errors.append(
                cls._build_error(
                    row=2,
                    column='Archivo',
                    value='',
                    error='El archivo no contiene registros para importar.',
                ),
            )

        return errors, rows_data

    @classmethod
    def _normalize_headers(cls, raw_headers):
        headers = []
        for header in raw_headers:
            if header is None:
                headers.append('')
            else:
                headers.append(str(header).strip())
        return headers

    @classmethod
    def _validar_headers(cls, headers):
        expected = set(cls.EXPECTED_HEADERS)
        present = {header for header in headers if header}
        missing = expected - present
        unexpected = present - expected - cls.OPTIONAL_HEADERS
        errors = []

        for header in sorted(missing):
            errors.append(
                cls._build_error(
                    row=1,
                    column=header,
                    value='',
                    error='Falta esta columna obligatoria en la plantilla.',
                ),
            )

        for header in sorted(unexpected):
            errors.append(
                cls._build_error(
                    row=1,
                    column=header,
                    value=header,
                    error='El encabezado no es válido para esta importación.',
                ),
            )

        if len(headers) < len(cls.EXPECTED_HEADERS):
            errors.append(
                cls._build_error(
                    row=1,
                    column='Archivo',
                    value='',
                    error=(
                        'La plantilla no contiene todas las columnas '
                        'esperadas.'
                    ),
                ),
            )

        return errors

    @classmethod
    def _validar_fila(
        cls,
        excel_row: int,
        row_data: Dict[str, Any],
        existing_products_by_internal_code,
        seen_barcodes,
        seen_internal_codes,
    ):
        errors = []
        normalized = {}
        codigo_interno = cls._parse_internal_code(
            row_data.get('Código Interno'),
            excel_row=excel_row,
            errors=errors,
        )
        existing_product = None
        if codigo_interno is not None:
            if codigo_interno in seen_internal_codes:
                errors.append(
                    cls._build_error(
                        row=excel_row,
                        column='Código Interno',
                        value=row_data.get('Código Interno'),
                        error='Este código interno está repetido en el archivo.',
                    ),
                )
            else:
                seen_internal_codes.add(codigo_interno)
                existing_product = existing_products_by_internal_code.get(
                    codigo_interno,
                )

        nombre = cls._parse_required_text(
            row_data.get('Nombre'),
            column='Nombre',
            excel_row=excel_row,
            errors=errors,
        )
        categoria_nombre = cls._parse_required_text(
            row_data.get('Categoría'),
            column='Categoría',
            excel_row=excel_row,
            errors=errors,
        )
        existencias = cls._parse_non_negative_integer(
            row_data.get('Existencias'),
            column='Existencias',
            excel_row=excel_row,
            errors=errors,
        )
        precio_compra = cls._parse_non_negative_decimal(
            row_data.get('Precio Compra'),
            column='Precio Compra',
            excel_row=excel_row,
            errors=errors,
        )
        precio_venta = cls._parse_non_negative_decimal(
            row_data.get('Precio Venta'),
            column='Precio Venta',
            excel_row=excel_row,
            errors=errors,
        )
        iva = cls._parse_iva(
            row_data.get('IVA (%)'),
            excel_row=excel_row,
            errors=errors,
        )
        fecha_ingreso = cls._parse_fecha(
            row_data.get('Fecha Ingreso'),
            column='Fecha Ingreso',
            excel_row=excel_row,
            errors=errors,
            default=timezone.now(),
            as_datetime=True,
        )
        fecha_caducidad = cls._parse_fecha(
            row_data.get('Fecha Caducidad'),
            column='Fecha Caducidad',
            excel_row=excel_row,
            errors=errors,
            default=None,
            as_datetime=False,
        )

        codigo_barras = cls._parse_optional_text(
            row_data.get('Código de Barras'),
        )
        marca = cls._parse_optional_text(row_data.get('Marca'))
        descripcion = cls._parse_optional_text(row_data.get('Descripción'))
        invima = cls._parse_optional_text(row_data.get('Invima'))

        if (
            precio_compra is not None
            and precio_venta is not None
            and precio_venta < precio_compra
        ):
            errors.append(
                cls._build_error(
                    row=excel_row,
                    column='Precio Venta',
                    value=row_data.get('Precio Venta'),
                    error='No puede ser menor que Precio Compra.',
                ),
            )

        if codigo_barras:
            existing_barcode_owner = Producto.objects.filter(
                empresa=get_empresa_actual_or_default(),
                codigo_barras=codigo_barras,
            ).first()
            if (
                existing_barcode_owner is not None
                and (
                    existing_product is None
                    or existing_barcode_owner.id != existing_product.id
                )
            ):
                errors.append(
                    cls._build_error(
                        row=excel_row,
                        column='Código de Barras',
                        value=codigo_barras,
                        error='Ya existe un producto con este código de barras.',
                    ),
                )
            elif codigo_barras in seen_barcodes:
                errors.append(
                    cls._build_error(
                        row=excel_row,
                        column='Código de Barras',
                        value=codigo_barras,
                        error='Este código de barras está repetido en el archivo.',
                    ),
                )
            else:
                seen_barcodes.add(codigo_barras)

        if errors:
            return errors, None

        normalized.update({
            'codigo_interno': codigo_interno,
            'existing_product_id': (
                existing_product.id if existing_product is not None else None
            ),
            'nombre': nombre,
            'categoria_nombre': categoria_nombre,
            'existencias': Decimal(existencias),
            'precio_compra': precio_compra,
            'precio_venta': precio_venta,
            'iva': iva,
            'fecha_ingreso': fecha_ingreso,
            'fecha_caducidad': fecha_caducidad,
            'codigo_barras': codigo_barras,
            'marca': marca,
            'descripcion': descripcion,
            'invima': invima,
            'unidad_medida_codigo': '94',
            'estandar_codigo': '999',
        })
        return errors, normalized

    @classmethod
    def _crear_productos(cls, rows_data, usuario=None):
        del usuario
        company = get_empresa_actual_or_default()
        categorias_cache = {
            categoria.nombre.upper(): categoria
            for categoria in Categoria.objects.filter(empresa=company)
        }

        for row_data in rows_data:
            categoria_key = row_data.pop('categoria_nombre').upper()
            existing_product_id = row_data.pop('existing_product_id', None)
            categoria = categorias_cache.get(categoria_key)
            if categoria is None:
                categoria = Categoria.objects.create(
                    empresa=company,
                    nombre=categoria_key,
                    descripcion='',
                )
                categorias_cache[categoria_key] = categoria

            if existing_product_id is not None:
                producto = Producto.objects.get(
                    id=existing_product_id,
                    empresa=company,
                )
                for campo, valor in row_data.items():
                    setattr(producto, campo, valor)
                producto.categoria = categoria
                producto.save(skip_full_clean=True)
            else:
                producto = Producto(
                    empresa=company,
                    categoria=categoria,
                    **row_data,
                )
                producto.save(skip_full_clean=True)
            producto.fecha_ingreso = row_data['fecha_ingreso']
            producto.save(update_fields=['fecha_ingreso', 'updated_at'])

    @staticmethod
    def _row_is_empty(row_values) -> bool:
        return all(
            value is None or str(value).strip() == ''
            for value in row_values
        )

    @classmethod
    def _parse_required_text(
        cls,
        value,
        column,
        excel_row,
        errors,
    ):
        text = cls._parse_optional_text(value)
        if not text:
            errors.append(
                cls._build_error(
                    row=excel_row,
                    column=column,
                    value=value,
                    error='Este campo es obligatorio.',
                ),
            )
            return None
        return text

    @staticmethod
    def _parse_optional_text(value) -> str:
        if value is None:
            return ''
        return str(value).strip()

    @classmethod
    def _parse_internal_code(
        cls,
        value,
        excel_row,
        errors,
    ):
        text = cls._parse_optional_text(value)
        if not text:
            return None

        try:
            decimal_value = Decimal(text)
        except Exception:
            errors.append(
                cls._build_error(
                    row=excel_row,
                    column='Código Interno',
                    value=value,
                    error='Debe ser un número entero válido.',
                ),
            )
            return None

        if decimal_value <= 0 or decimal_value != decimal_value.to_integral():
            errors.append(
                cls._build_error(
                    row=excel_row,
                    column='Código Interno',
                    value=value,
                    error='Debe ser un número entero válido.',
                ),
            )
            return None

        return int(decimal_value)

    @classmethod
    def _parse_non_negative_integer(
        cls,
        value,
        column,
        excel_row,
        errors,
    ):
        if value in (None, ''):
            errors.append(
                cls._build_error(
                    row=excel_row,
                    column=column,
                    value=value,
                    error='Este campo es obligatorio.',
                ),
            )
            return None

        try:
            decimal_value = Decimal(str(value))
        except Exception:
            errors.append(
                cls._build_error(
                    row=excel_row,
                    column=column,
                    value=value,
                    error='Debe ser un número entero mayor o igual a 0.',
                ),
            )
            return None

        if decimal_value < 0 or decimal_value != decimal_value.to_integral():
            errors.append(
                cls._build_error(
                    row=excel_row,
                    column=column,
                    value=value,
                    error='Debe ser un número entero mayor o igual a 0.',
                ),
            )
            return None

        return int(decimal_value)

    @classmethod
    def _parse_non_negative_decimal(
        cls,
        value,
        column,
        excel_row,
        errors,
    ):
        if value in (None, ''):
            errors.append(
                cls._build_error(
                    row=excel_row,
                    column=column,
                    value=value,
                    error='Este campo es obligatorio.',
                ),
            )
            return None

        try:
            decimal_value = Decimal(str(value))
        except Exception:
            errors.append(
                cls._build_error(
                    row=excel_row,
                    column=column,
                    value=value,
                    error='Debe ser un número mayor o igual a 0.',
                ),
            )
            return None

        if decimal_value < 0:
            errors.append(
                cls._build_error(
                    row=excel_row,
                    column=column,
                    value=value,
                    error='Debe ser un número mayor o igual a 0.',
                ),
            )
            return None

        return decimal_value.quantize(Decimal('0.01'))

    @classmethod
    def _parse_iva(
        cls,
        value,
        excel_row,
        errors,
    ):
        if value in (None, ''):
            return Decimal('0.00')

        try:
            iva = Decimal(str(value))
        except Exception:
            errors.append(
                cls._build_error(
                    row=excel_row,
                    column='IVA (%)',
                    value=value,
                    error='Debe ser un número entre 0 y 100.',
                ),
            )
            return None

        if iva < 0 or iva > 100:
            errors.append(
                cls._build_error(
                    row=excel_row,
                    column='IVA (%)',
                    value=value,
                    error='Debe ser un número entre 0 y 100.',
                ),
            )
            return None

        return iva.quantize(Decimal('0.01'))

    @classmethod
    def _parse_fecha(
        cls,
        value,
        column,
        excel_row,
        errors,
        default,
        as_datetime,
    ):
        if value in (None, ''):
            return default

        parsed = None
        if isinstance(value, datetime):
            parsed = value
        elif isinstance(value, date):
            parsed = datetime.combine(value, datetime.min.time())
        else:
            value_str = str(value).strip()
            formatos = (
                '%Y-%m-%d',
                '%d/%m/%Y',
                '%d-%m-%Y',
                '%Y/%m/%d',
            )
            for formato in formatos:
                try:
                    parsed = datetime.strptime(value_str, formato)
                    break
                except ValueError:
                    continue

        if parsed is None:
            errors.append(
                cls._build_error(
                    row=excel_row,
                    column=column,
                    value=value,
                    error='Debe ser una fecha válida.',
                ),
            )
            return None

        if as_datetime:
            return timezone.make_aware(
                parsed,
                timezone.get_current_timezone(),
            ) if timezone.is_naive(parsed) else parsed
        return parsed.date()

    @staticmethod
    def _build_error(row, column, value, error):
        return {
            'row': row,
            'column': column,
            'value': '' if value is None else str(value),
            'error': error,
        }


class StockService:
    """
    Servicio para gestionar la lógica de negocio de stock e inventario.

    Responsable de actualizar existencias, validar disponibilidad
    y registrar movimientos en el historial de inventario.
    """

    @staticmethod
    @transaction.atomic
    def actualizar_stock(
        producto_id: int,
        cantidad: Decimal,
        tipo: str,
        motivo: str,
        usuario: Usuario,
        precio_unitario: Optional[Decimal] = None,
        factura: Optional[FacturaCompra] = None,
        venta=None,
        observaciones: str = '',
    ) -> HistorialInventario:
        """
        Actualiza el stock de un producto y registra el movimiento.

        Args:
            producto_id: ID del producto
            cantidad: Cantidad (positiva para entrada, negativa para salida)
            tipo: Tipo de movimiento (ENTRADA, SALIDA, AJUSTE)
            motivo: Motivo del movimiento
            usuario: Usuario que realiza el movimiento
            precio_unitario: Precio unitario al momento del movimiento
            factura: Factura de compra asociada (opcional)
            venta: Venta asociada (opcional)
            observaciones: Observaciones adicionales (opcional)

        Returns:
            HistorialInventario: Registro del movimiento creado

        Raises:
            ProductoNoEncontradoError: Si el producto no existe
            StockInsuficienteError: Si no hay stock suficiente para salida
        """
        producto = ProductoService.obtener_producto(producto_id)

        if tipo == HistorialInventario.TIPO_SALIDA and cantidad > 0:
            cantidad = -cantidad

        if cantidad < 0:
            try:
                producto.actualizar_stock(cantidad)
            except ValueError:
                raise StockInsuficienteError(
                    producto.nombre,
                    producto.existencias,
                    -cantidad,
                )
        else:
            producto.actualizar_stock(cantidad)

        if precio_unitario is None:
            precio_unitario = producto.precio_compra

        historial = HistorialInventario.objects.create(
            empresa=producto.empresa,
            producto=producto,
            tipo_movimiento=tipo,
            cantidad=cantidad,
            precio_unitario=precio_unitario,
            factura=factura,
            venta=venta,
            motivo=motivo,
            usuario=usuario,
            observaciones=observaciones,
        )

        return historial

    @staticmethod
    def validar_disponibilidad(
        producto_id: int,
        cantidad: Decimal
    ) -> bool:
        """
        Verifica si hay suficiente stock disponible.

        Args:
            producto_id: ID del producto
            cantidad: Cantidad requerida (positiva)

        Returns:
            bool: True si hay suficiente stock

        Raises:
            ProductoNoEncontradoError: Si el producto no existe
            StockInsuficienteError: Si no hay stock suficiente
        """
        producto = ProductoService.obtener_producto(producto_id)

        if not producto.validar_stock(cantidad):
            raise StockInsuficienteError(
                producto.nombre,
                producto.existencias,
                cantidad,
            )

        return True

    @staticmethod
    @transaction.atomic
    def ajustar_inventario(
        producto_id: int,
        nueva_cantidad: Decimal,
        motivo: str,
        usuario: Usuario,
        observaciones: str = '',
    ) -> HistorialInventario:
        """
        Realiza un ajuste manual de inventario.

        Establece las existencias a un valor específico y registra
        la diferencia como un movimiento de tipo AJUSTE.

        Args:
            producto_id: ID del producto
            nueva_cantidad: Nueva cantidad de existencias
            motivo: Motivo del ajuste
            usuario: Usuario que realiza el ajuste
            observaciones: Observaciones adicionales (opcional)

        Returns:
            HistorialInventario: Registro del ajuste creado
        """
        producto = ProductoService.obtener_producto(producto_id)
        diferencia = nueva_cantidad - producto.existencias

        historial = HistorialInventario.objects.create(
            empresa=producto.empresa,
            producto=producto,
            tipo_movimiento=HistorialInventario.TIPO_AJUSTE,
            cantidad=diferencia,
            precio_unitario=producto.precio_compra,
            motivo=motivo,
            usuario=usuario,
            observaciones=observaciones,
        )

        producto.existencias = nueva_cantidad
        producto.save(update_fields=['existencias', 'updated_at'])

        return historial


class FacturaCompraService:
    """
    Servicio para gestionar la lógica de negocio de facturas de compra.

    Maneja el registro, procesamiento y consulta de facturas,
    incluyendo la actualización automática de inventario.
    """

    @staticmethod
    @transaction.atomic
    def registrar_factura_compra(data: Dict[str, Any]) -> FacturaCompra:
        """
        Registra una nueva factura de compra con sus detalles.

        Crea la factura y sus detalles, luego calcula los totales.
        NO actualiza el inventario automáticamente (usar procesar_factura).
        Si no se proporciona proveedor, se asigna uno genérico por defecto.

        Args:
            data: Datos de la factura incluyendo 'detalles'

        Returns:
            FacturaCompra: Instancia de la factura creada

        Raises:
            FacturaSinDetallesError: Si no se incluyen detalles
            ProductoNoEncontradoError: Si un producto no existe
        """
        detalles_data = data.pop('detalles', [])
        empresa = get_empresa_actual_or_default()
        data['empresa'] = empresa

        if not detalles_data:
            raise FacturaSinDetallesError(0)

        if 'proveedor' not in data or data['proveedor'] is None:
            proveedor_default, _ = Proveedor.objects.get_or_create(
                empresa=empresa,
                numero_documento='0000000000',
                defaults={
                    'razon_social': 'PROVEEDOR GENERAL',
                    'nombre_contacto': 'SIN CONTACTO',
                    'email': 'proveedor@default.com',
                    'telefono': '0000000000',
                    'direccion': 'SIN DIRECCION',
                    'ciudad': 'SIN CIUDAD',
                    'departamento': 'SIN DEPARTAMENTO',
                    'tipo_productos': 'GENERAL',
                }
            )
            data['proveedor'] = proveedor_default

        factura = FacturaCompra.objects.create(**data)

        for detalle_data in detalles_data:
            DetalleFacturaCompra.objects.create(
                factura=factura,
                empresa=empresa,
                **detalle_data
            )

        factura.calcular_totales()
        factura.save(update_fields=['subtotal', 'iva', 'total', 'updated_at'])

        return FacturaCompra.objects.prefetch_related(
            'detalles__producto__categoria'
        ).select_related(
            'proveedor', 'usuario_registro'
        ).get(id=factura.id, empresa=empresa)

    @staticmethod
    @transaction.atomic
    def procesar_factura(
        factura_id: int,
        usuario: Usuario,
        pricing_rules: Optional[Dict[str, Any]] = None,
    ) -> FacturaCompra:
        """
        Procesa una factura de compra y actualiza el inventario.

        Por cada detalle de la factura, crea un movimiento de entrada
        en el historial y actualiza el stock del producto.
        La factura pasa a estado PROCESADA.

        Args:
            factura_id: ID de la factura a procesar
            usuario: Usuario que procesa la factura

        Returns:
            FacturaCompra: Factura procesada

        Raises:
            FacturaNoEncontradaError: Si la factura no existe
            FacturaYaProcesadaError: Si la factura ya fue procesada
            FacturaSinDetallesError: Si la factura no tiene detalles
        """
        try:
            factura = FacturaCompra.objects.prefetch_related(
                'detalles__producto'
            ).get(id=factura_id, empresa=get_empresa_actual_or_default())
        except FacturaCompra.DoesNotExist:
            raise FacturaNoEncontradaError(factura_id)

        if factura.estado == FacturaCompra.ESTADO_PROCESADA:
            raise FacturaYaProcesadaError(factura_id)

        detalles = factura.detalles.all()
        if not detalles:
            raise FacturaSinDetallesError(factura_id)

        for detalle in detalles:
            StockService.actualizar_stock(
                producto_id=detalle.producto_id,
                cantidad=detalle.cantidad,
                tipo=HistorialInventario.TIPO_ENTRADA,
                motivo=_("Entrada por factura de compra: %(num)s") % {
                    'num': factura.numero_factura
                },
                usuario=usuario,
                precio_unitario=detalle.precio_unitario,
                factura=factura,
                observaciones=_(
                    "Procesamiento automático de factura. "
                    "Producto: %(producto)s, Cantidad: %(cantidad)s, "
                    "Precio: %(precio)s"
                ) % {
                    'producto': detalle.producto.nombre,
                    'cantidad': detalle.cantidad,
                    'precio': detalle.precio_unitario,
                },
            )

            producto = detalle.producto
            if detalle.precio_unitario > 0:
                q = Decimal('0.01')
                iva_decimal = detalle.iva / Decimal('100')
                costo_final = (
                    detalle.precio_unitario * (Decimal('1') + iva_decimal)
                ).quantize(q)
                producto.precio_compra = detalle.precio_unitario.quantize(q)
                if detalle.precio_venta_sugerido:
                    producto.precio_venta = (
                        detalle.precio_venta_sugerido
                    ).quantize(q)
                else:
                    producto.precio_venta = calculate_suggested_sale_price(
                        costo_final,
                        pricing_rules=pricing_rules,
                    )
                producto.iva = detalle.iva
                producto.save(update_fields=[
                    'precio_compra', 'precio_venta', 'iva', 'updated_at'
                ])

        factura.marcar_como_procesada()

        return FacturaCompra.objects.prefetch_related(
            'detalles__producto__categoria',
            'movimientos_inventario',
        ).select_related(
            'proveedor', 'usuario_registro'
        ).get(id=factura.id, empresa=get_empresa_actual_or_default())

    @staticmethod
    def obtener_factura(factura_id: int) -> FacturaCompra:
        """
        Obtiene una factura de compra con todos sus detalles.

        Args:
            factura_id: ID de la factura

        Returns:
            FacturaCompra: Factura con relaciones precargadas

        Raises:
            FacturaNoEncontradaError: Si la factura no existe
        """
        try:
            return FacturaCompra.objects.prefetch_related(
                'detalles__producto__categoria',
                'movimientos_inventario',
            ).select_related(
                'proveedor', 'usuario_registro'
            ).get(id=factura_id, empresa=get_empresa_actual_or_default())
        except FacturaCompra.DoesNotExist:
            raise FacturaNoEncontradaError(factura_id)

    @staticmethod
    def listar_facturas(
        filtros: Optional[Dict[str, Any]] = None
    ) -> List[FacturaCompra]:
        """
        Lista facturas de compra con filtros opcionales.

        Args:
            filtros: Diccionario con filtros (proveedor_id, estado,
                     fecha_desde, fecha_hasta, q)

        Returns:
            List[FacturaCompra]: Lista de facturas filtradas
        """
        queryset = FacturaCompra.objects.select_related(
            'proveedor', 'usuario_registro'
        ).prefetch_related('detalles').filter(
            empresa=get_empresa_actual_or_default(),
        )

        if not filtros:
            return list(queryset.order_by('-fecha_registro'))

        q_objects = Q()

        if filtros.get('proveedor_id'):
            q_objects &= Q(proveedor_id=filtros['proveedor_id'])

        if filtros.get('estado'):
            q_objects &= Q(estado=filtros['estado'])

        if filtros.get('fecha_desde'):
            q_objects &= Q(fecha_factura__gte=filtros['fecha_desde'])

        if filtros.get('fecha_hasta'):
            q_objects &= Q(fecha_factura__lte=filtros['fecha_hasta'])

        if filtros.get('q'):
            q = filtros['q']
            q_objects &= (
                Q(numero_factura__icontains=q) |
                Q(proveedor__razon_social__icontains=q) |
                Q(observaciones__icontains=q)
            )

        if q_objects:
            queryset = queryset.filter(q_objects)

        return list(queryset.order_by('-fecha_registro'))


class ReporteService:
    """
    Servicio para generar reportes y estadísticas de inventario.
    """

    @staticmethod
    def calcular_valor_total_inventario() -> Dict[str, Decimal]:
        """
        Calcula el valor total del inventario.

        Returns:
            dict: Diccionario con valor_compra, valor_venta,
                  cantidad_productos y total_existencias
        """
        productos = Producto.objects.filter(
            empresa=get_empresa_actual_or_default()
        )
        total_valor_compra = Decimal('0.00')
        total_valor_venta = Decimal('0.00')
        total_existencias = Decimal('0')
        cantidad_productos = 0

        for p in productos:
            total_valor_compra += p.calcular_valor_inventario()
            total_valor_venta += p.calcular_valor_venta()
            total_existencias += p.existencias if p.existencias else Decimal('0')
            cantidad_productos += 1

        return {
            'valor_compra': total_valor_compra,
            'valor_venta': total_valor_venta,
            'cantidad_productos': cantidad_productos,
            'total_existencias': total_existencias,
        }

    @staticmethod
    def productos_bajo_stock() -> List[Producto]:
        """
        Retorna productos con existencias por debajo del mínimo.

        Args:
            minimo: Cantidad mínima de stock (default: 10)

        Returns:
            List[Producto]: Productos con stock bajo
        """
        return list(
            Producto.objects.select_related('categoria')
            .filter(
                empresa=get_empresa_actual_or_default(),
                existencias__lte=F('stock_minimo'),
            )
            .order_by('existencias')
        )

    @staticmethod
    def productos_mas_vendidos(
        fecha_inicio: Optional[date] = None,
        fecha_fin: Optional[date] = None,
        limite: int = 10,
    ) -> List[Dict]:
        """
        Retorna los productos más vendidos en un período.

        Requiere que el módulo ventas esté instalado para obtener
        los datos de ventas reales.

        Args:
            fecha_inicio: Fecha de inicio del período
            fecha_fin: Fecha de fin del período
            limite: Cantidad máxima de productos (default: 10)

        Returns:
            List[Dict]: Lista de productos con total_vendido
        """
        from django.apps import apps

        if not apps.is_installed('ventas'):
            return ReporteService.productos_bajo_stock()[:limite]

        try:
            Venta = apps.get_model('ventas', 'Venta')
            DetalleVenta = apps.get_model('ventas', 'DetalleVenta')
        except LookupError:
            return ReporteService.productos_bajo_stock()[:limite]

        ventas_qs = Venta.objects.filter(
            empresa=get_empresa_actual_or_default(),
            estado__in=['TERMINADA', 'PENDIENTE']
        )

        if fecha_inicio:
            ventas_qs = ventas_qs.filter(fecha_venta__date__gte=fecha_inicio)
        if fecha_fin:
            ventas_qs = ventas_qs.filter(fecha_venta__date__lte=fecha_fin)

        resultados = (
            DetalleVenta.objects.filter(
                venta__in=ventas_qs
            )
            .values('producto_id', 'producto__nombre', 'producto__codigo_interno')
            .annotate(
                total_vendido=Sum('cantidad', output_field=DecimalField())
            )
            .order_by('-total_vendido')[:limite]
        )

        return [
            {
                'producto_id': r['producto_id'],
                'nombre': r['producto__nombre'],
                'codigo_interno': r['producto__codigo_interno'],
                'total_vendido': r['total_vendido'] or Decimal('0'),
            }
            for r in resultados
        ]


class HistorialService:
    """
    Servicio para consultar el historial de movimientos de inventario.
    """

    @staticmethod
    def obtener_historial_producto(
        producto_id: int,
        filtros: Optional[Dict[str, Any]] = None
    ) -> List[HistorialInventario]:
        """
        Obtiene los movimientos de inventario de un producto específico.

        Args:
            producto_id: ID del producto
            filtros: Filtros opcionales (tipo_movimiento, fecha_desde,
                     fecha_hasta, limite)

        Returns:
            List[HistorialInventario]: Lista de movimientos

        Raises:
            ProductoNoEncontradoError: Si el producto no existe
        """
        producto = ProductoService.obtener_producto(producto_id)
        queryset = HistorialInventario.objects.filter(producto=producto)

        if filtros:
            if filtros.get('tipo_movimiento'):
                queryset = queryset.filter(
                    tipo_movimiento=filtros['tipo_movimiento']
                )

            if filtros.get('fecha_desde'):
                queryset = queryset.filter(
                    fecha__gte=filtros['fecha_desde']
                )

            if filtros.get('fecha_hasta'):
                queryset = queryset.filter(
                    fecha__lte=filtros['fecha_hasta']
                )

        limite = filtros.get('limite', 100) if filtros else 100
        return list(
            queryset.select_related(
                'producto', 'usuario', 'factura'
            ).order_by('-fecha')[:limite]
        )

    @staticmethod
    def obtener_historial_general(
        filtros: Optional[Dict[str, Any]] = None
    ) -> List[HistorialInventario]:
        """
        Obtiene el historial completo de movimientos de inventario.

        Args:
            filtros: Filtros opcionales (producto_id, tipo_movimiento,
                     usuario_id, fecha_desde, fecha_hasta, limite)

        Returns:
            List[HistorialInventario]: Lista de movimientos filtrados
        """
        queryset = HistorialInventario.objects.select_related(
            'producto', 'usuario', 'factura', 'venta'
        )

        if not filtros:
            return list(queryset.order_by('-fecha')[:100])

        q_objects = Q()

        if filtros.get('producto_id'):
            q_objects &= Q(producto_id=filtros['producto_id'])

        if filtros.get('tipo_movimiento'):
            q_objects &= Q(tipo_movimiento=filtros['tipo_movimiento'])

        if filtros.get('usuario_id'):
            q_objects &= Q(usuario_id=filtros['usuario_id'])

        if filtros.get('fecha_desde'):
            q_objects &= Q(fecha__gte=filtros['fecha_desde'])

        if filtros.get('fecha_hasta'):
            q_objects &= Q(fecha__lte=filtros['fecha_hasta'])

        if filtros.get('motivo'):
            q_objects &= Q(motivo__icontains=filtros['motivo'])

        if q_objects:
            queryset = queryset.filter(q_objects)

        limite = filtros.get('limite', 100)
        return list(queryset.order_by('-fecha')[:limite])
